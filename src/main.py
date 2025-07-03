# 1. Imports
import json
import os
import shutil
import time
import re
import urllib.parse
import calendar
import difflib
import csv
import base64
import sqlite3
from datetime import datetime, timedelta
from colorama import Fore, Back, Style, init
from tqdm import tqdm
import getpass
import hashlib
try:
    from .integrity_checker import DataIntegrityChecker
    from .data_compression import DataCompressor
except ImportError:
    from integrity_checker import DataIntegrityChecker
    from data_compression import DataCompressor

# Cryptography imports for password encryption
try:
    from cryptography.fernet import Fernet
    from cryptography.hazmat.primitives import hashes
    from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
    CRYPTOGRAPHY_AVAILABLE = True
except ImportError:
    CRYPTOGRAPHY_AVAILABLE = False
    print("Warning: cryptography library not available. Passwords will be stored in plain text.")
    print("Install with: pip install cryptography")

# Initialize colorama for Windows compatibility
init(autoreset=True)

# 2. Configuration
BILLS_FILE = 'bills.json'  # Legacy JSON file (for migration)
BACKUP_DIR = 'backups'
MAX_BACKUPS = 5
DATE_FORMAT = '%Y-%m-%d'
TEMPLATES_FILE = 'bill_templates.json'  # Legacy JSON file (for migration)

# Database configuration
DB_FILE = 'bills_tracker.db'

# Encryption configuration
ENCRYPTION_KEY_FILE = '.encryption_key'
SALT_FILE = '.salt'
MASTER_PASSWORD_FILE = '.master_password'

# Session timeout configuration
SESSION_TIMEOUT_MINUTES = 30  # Auto-exit after 30 minutes of inactivity
SESSION_CONFIG_FILE = '.session_config'

bills = []
templates = []
bill_templates = []

# Global session variables
session_start_time = None
last_activity_time = None
session_locked = False

# 3. Color utility functions
class Colors:
    """Color constants and helper functions."""
    
    # Status colors
    SUCCESS = Fore.GREEN
    ERROR = Fore.RED
    WARNING = Fore.YELLOW
    INFO = Fore.CYAN
    TITLE = Fore.MAGENTA
    
    # Bill status colors
    PAID = Fore.GREEN
    UNPAID = Fore.YELLOW
    OVERDUE = Fore.RED + Style.BRIGHT
    DUE_SOON = Fore.YELLOW + Style.BRIGHT
    
    # UI colors
    MENU = Fore.BLUE + Style.BRIGHT
    PROMPT = Fore.WHITE + Style.BRIGHT
    RESET = Style.RESET_ALL

def colored_print(text, color=Colors.RESET):
    """Print text with color."""
    print(f"{color}{text}{Colors.RESET}")

def colored_input(prompt, color=Colors.PROMPT):
    """Get input with colored prompt and session timeout check."""
    # Check timeout before accepting input
    if check_session_timeout():
        # Session timeout will exit the app
        pass
    # Get input and update activity
    user_input = input(f"{color}{prompt}{Colors.RESET}")
    update_activity()
    return user_input

# 4.2 Encryption utility functions
class PasswordEncryption:
    """Password encryption and decryption utilities using Fernet."""
    
    def __init__(self):
        self.fernet = None
        self.key = None
        self.salt = None
    
    def generate_salt(self):
        """Generate a random salt for key derivation."""
        if not CRYPTOGRAPHY_AVAILABLE:
            return None
        return os.urandom(16)
    
    def derive_key_from_password(self, password, salt):
        """Derive encryption key from password and salt."""
        if not CRYPTOGRAPHY_AVAILABLE:
            return None
        
        kdf = PBKDF2HMAC(
            algorithm=hashes.SHA256(),
            length=32,
            salt=salt,
            iterations=100000,
        )
        key = base64.urlsafe_b64encode(kdf.derive(password.encode()))
        return key
    
    def initialize_encryption(self, master_password=None):
        """Initialize encryption with master password or generate new key."""
        if not CRYPTOGRAPHY_AVAILABLE:
            return False
        
        try:
            # Check if key file exists
            if os.path.exists(ENCRYPTION_KEY_FILE) and os.path.exists(SALT_FILE):
                # Load existing key
                with open(SALT_FILE, 'rb') as f:
                    self.salt = f.read()
                
                if master_password:
                    # Derive key from provided password
                    self.key = self.derive_key_from_password(master_password, self.salt)
                else:
                    # Load stored key
                    with open(ENCRYPTION_KEY_FILE, 'rb') as f:
                        self.key = f.read()
            else:
                # Generate new key
                if master_password:
                    # Use provided password
                    self.salt = self.generate_salt()
                    self.key = self.derive_key_from_password(master_password, self.salt)
                else:
                    # Generate random key
                    self.key = Fernet.generate_key()
                    self.salt = self.generate_salt()
                
                # Save key and salt
                with open(ENCRYPTION_KEY_FILE, 'wb') as f:
                    f.write(self.key)
                with open(SALT_FILE, 'wb') as f:
                    f.write(self.salt)
            
            self.fernet = Fernet(self.key)
            return True
            
        except Exception as e:
            error_msg(f"Failed to initialize encryption: {e}")
            return False
    
    def encrypt_password(self, password):
        """Encrypt a password."""
        if not CRYPTOGRAPHY_AVAILABLE or not self.fernet or not password:
            return password
        
        try:
            encrypted = self.fernet.encrypt(password.encode())
            return base64.urlsafe_b64encode(encrypted).decode()
        except Exception as e:
            error_msg(f"Failed to encrypt password: {e}")
            return password
    
    def decrypt_password(self, encrypted_password):
        """Decrypt a password."""
        if not CRYPTOGRAPHY_AVAILABLE or not self.fernet or not encrypted_password:
            return encrypted_password
        
        try:
            # Check if password is already encrypted
            if encrypted_password.startswith('gAAAAA'):
                # This is a Fernet token, decrypt it
                encrypted_bytes = base64.urlsafe_b64decode(encrypted_password.encode())
                decrypted = self.fernet.decrypt(encrypted_bytes)
                return decrypted.decode()
            else:
                # This might be a plain text password, return as is
                return encrypted_password
        except Exception as e:
            # If decryption fails, assume it's plain text
            return encrypted_password
    
    def migrate_passwords(self, bills_data):
        """Migrate plain text passwords to encrypted format."""
        if not CRYPTOGRAPHY_AVAILABLE or not self.fernet:
            return bills_data
        
        migrated = False
        for bill in bills_data:
            if 'password' in bill and bill['password']:
                # Check if password is already encrypted
                if not bill['password'].startswith('gAAAAA'):
                    # Encrypt plain text password
                    bill['password'] = self.encrypt_password(bill['password'])
                    migrated = True
        
        if migrated:
            success_msg("Passwords migrated to encrypted format")
        
        return bills_data

# Global encryption instance
password_encryption = PasswordEncryption()

# Database functions
def get_db_connection():
    """Get a connection to the SQLite database."""
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn

def initialize_database():
    """Initialize the SQLite database with required tables."""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Create bills table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS bills (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            due_date TEXT NOT NULL,
            billing_cycle TEXT,
            reminder_days INTEGER,
            web_page TEXT,
            login_info TEXT,
            password TEXT,
            paid INTEGER DEFAULT 0,
            company_email TEXT,
            support_phone TEXT,
            billing_phone TEXT,
            customer_service_hours TEXT,
            account_number TEXT,
            reference_id TEXT,
            support_chat_url TEXT,
            mobile_app TEXT
        )
    ''')
    
    # Create templates table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS templates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            due_date TEXT,
            billing_cycle TEXT,
            reminder_days INTEGER,
            web_page TEXT,
            login_info TEXT,
            password TEXT,
            company_email TEXT,
            support_phone TEXT,
            billing_phone TEXT,
            customer_service_hours TEXT,
            account_number TEXT,
            reference_id TEXT,
            support_chat_url TEXT,
            mobile_app TEXT
        )
    ''')
    
    conn.commit()
    conn.close()

def check_session_timeout():
    """Check if session has timed out."""
    if last_activity_time is None:
        return False
    
    if (datetime.now() - last_activity_time) > timedelta(minutes=SESSION_TIMEOUT_MINUTES):
        print(f"\n🔒 Session expired due to {SESSION_TIMEOUT_MINUTES} minutes of inactivity.")
        print("🔄 Exiting application for security...")
        success_msg("Thank you for using Bills Tracker! 👋")
        os._exit(0)  # Force exit immediately
        return True
    return False

def start_session():
    """Start a new session and initialize activity tracking."""
    global session_start_time, last_activity_time, session_locked
    session_start_time = datetime.now()
    last_activity_time = datetime.now()
    session_locked = False

def update_activity():
    """Update the last activity time to prevent session timeout."""
    global last_activity_time
    last_activity_time = datetime.now()

def success_msg(message):
    """Print success message."""
    colored_print(f"✅ {message}", Colors.SUCCESS)

def error_msg(message):
    """Print error message."""
    colored_print(f"❌ {message}", Colors.ERROR)

def warning_msg(message):
    """Print warning message."""
    colored_print(f"⚠️  {message}", Colors.WARNING)

def info_msg(message):
    """Print info message."""
    colored_print(f"ℹ️  {message}", Colors.INFO)

def title_msg(message):
    """Print title message."""
    colored_print(f"🏠 {message}", Colors.TITLE + Style.BRIGHT)

# 4. Utility functions
def clear_console():
    """Clear the console screen."""
    os.system('cls' if os.name == 'nt' else 'clear')

# 4.1 Auto-complete functions
class AutoComplete:
    """Auto-complete functionality for bill names and other fields."""
    
    @staticmethod
    def get_bill_names():
        """Get all bill names for auto-completion."""
        return [bill['name'] for bill in bills]
    
    @staticmethod
    def get_websites():
        """Get all unique websites for auto-completion."""
        websites = set()
        for bill in bills:
            if bill.get('web_page'):
                websites.add(bill['web_page'])
        return list(websites)
    
    @staticmethod
    def suggest_names(partial_input, max_suggestions=5):
        """Suggest bill names based on partial input."""
        if not partial_input or not bills:
            return []
        
        bill_names = AutoComplete.get_bill_names()
        partial_lower = partial_input.lower()
        
        # Find exact matches first
        exact_matches = [name for name in bill_names if name.lower().startswith(partial_lower)]
        
        # Find fuzzy matches if we need more suggestions
        if len(exact_matches) < max_suggestions:
            fuzzy_matches = difflib.get_close_matches(
                partial_input, 
                bill_names, 
                n=max_suggestions - len(exact_matches),
                cutoff=0.3
            )
            # Remove duplicates
            fuzzy_matches = [name for name in fuzzy_matches if name not in exact_matches]
            exact_matches.extend(fuzzy_matches)
        
        return exact_matches[:max_suggestions]
    
    @staticmethod
    def suggest_websites(partial_input, max_suggestions=3):
        """Suggest websites based on partial input."""
        if not partial_input or not bills:
            return []
        
        websites = AutoComplete.get_websites()
        partial_lower = partial_input.lower()
        
        # Find matches
        matches = [site for site in websites if partial_lower in site.lower()]
        return matches[:max_suggestions]

def get_input_with_autocomplete(prompt, autocomplete_type="bills", allow_empty=False):
    """Get input with auto-complete suggestions."""
    if autocomplete_type not in ["bills", "websites"]:
        autocomplete_type = "bills"
    
    print(f"{Colors.PROMPT}{prompt}{Colors.RESET}")
    if autocomplete_type == "bills" and bills:
        print(f"{Colors.INFO}💡 Type to see suggestions (available bills: {len(bills)}){Colors.RESET}")
    elif autocomplete_type == "websites":
        websites = AutoComplete.get_websites()
        if websites:
            print(f"{Colors.INFO}💡 Type to see website suggestions (available: {len(websites)}){Colors.RESET}")
    
    user_input = ""
    suggestions_shown = False
    
    while True:
        if not suggestions_shown:
            current_input = input(f"{Colors.PROMPT}> {Colors.RESET}").strip()
        else:
            current_input = input(f"{Colors.PROMPT}> {user_input}{Colors.RESET}").strip()
        
        # Handle special commands
        if current_input.lower() == 'cancel':
            return None
        elif current_input.lower() == 'help':
            show_autocomplete_help(autocomplete_type)
            continue
        elif current_input.lower().startswith('?'):
            # Show all available options
            show_all_options(autocomplete_type)
            continue
        elif current_input == "":
            if allow_empty:
                return ""
            if user_input:
                return user_input
            continue
        
        # Update user input
        if not suggestions_shown:
            user_input = current_input
        else:
            user_input += current_input
        
        # Get suggestions
        if autocomplete_type == "bills":
            suggestions = AutoComplete.suggest_names(user_input)
        else:
            suggestions = AutoComplete.suggest_websites(user_input)
        
        if suggestions and len(user_input) > 0:
            print(f"\n{Colors.SUCCESS}💡 Suggestions:{Colors.RESET}")
            for i, suggestion in enumerate(suggestions, 1):
                # Highlight the matching part
                if user_input.lower() in suggestion.lower():
                    highlighted = suggestion.replace(
                        user_input, 
                        f"{Colors.WARNING}{user_input}{Colors.RESET}{Colors.INFO}"
                    )
                    print(f"{Colors.INFO}  {i}. {highlighted}{Colors.RESET}")
                else:
                    print(f"{Colors.INFO}  {i}. {suggestion}{Colors.RESET}")
            
            print(f"{Colors.INFO}  0. Continue typing...{Colors.RESET}")
            print(f"{Colors.INFO}  ?. Show all options{Colors.RESET}")
            print(f"{Colors.WARNING}  Enter number to select, or continue typing{Colors.RESET}")
            
            choice = input(f"{Colors.PROMPT}Choice (or continue typing): {Colors.RESET}").strip()
            
            if choice.isdigit():
                choice_num = int(choice)
                if 1 <= choice_num <= len(suggestions):
                    return suggestions[choice_num - 1]
                elif choice_num == 0:
                    suggestions_shown = True
                    continue
            elif choice.lower() == 'cancel':
                return None
            elif choice.lower().startswith('?'):
                show_all_options(autocomplete_type)
                continue
            else:
                # User wants to continue typing
                user_input += choice
                suggestions_shown = True
                continue
        else:
            # No suggestions, accept input or ask to continue
            if user_input:
                confirm = input(f"{Colors.WARNING}No suggestions found. Use '{user_input}'? (y/n/continue): {Colors.RESET}").strip().lower()
                if confirm in ['y', 'yes']:
                    return user_input
                elif confirm in ['n', 'no']:
                    user_input = ""
                    suggestions_shown = False
                    continue
                else:
                    suggestions_shown = True
                    continue
            else:
                continue

def show_autocomplete_help(autocomplete_type):
    """Show help for auto-complete feature."""
    print(f"\n{Colors.TITLE}📚 Auto-Complete Help{Colors.RESET}")
    print(f"{Colors.INFO}Available commands:{Colors.RESET}")
    print(f"  • Type partial name to see suggestions")
    print(f"  • Enter number to select a suggestion")
    print(f"  • Type '?' to see all available options")
    print(f"  • Type 'cancel' to cancel input")
    print(f"  • Type 'help' to see this help")
    print(f"  • Press Enter with empty input to continue\n")

def show_all_options(autocomplete_type):
    """Show all available options for auto-complete."""
    if autocomplete_type == "bills":
        if not bills:
            warning_msg("No bills available yet.")
            return
        
        print(f"\n{Colors.TITLE}📋 All Available Bills:{Colors.RESET}")
        for i, bill in enumerate(bills, 1):
            status = "✓ Paid" if bill.get('paid', False) else "○ Unpaid"
            print(f"{Colors.INFO}  {i:2}. {bill['name']} [{status}] - Due: {bill['due_date']}{Colors.RESET}")
    
    elif autocomplete_type == "websites":
        websites = AutoComplete.get_websites()
        if not websites:
            warning_msg("No websites available yet.")
            return
        
        print(f"\n{Colors.TITLE}🌐 All Available Websites:{Colors.RESET}")
        for i, website in enumerate(websites, 1):
            print(f"{Colors.INFO}  {i:2}. {website}{Colors.RESET}")
    
    print()

# 5. File operations
def load_bills():
    """Load bills from SQLite database."""
    global bills
    try:
        # Initialize database if it doesn't exist
        initialize_database()
        
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM bills ORDER BY due_date')
        rows = cursor.fetchall()
        
        bills = []
        for row in rows:
            bill = dict(row)
            # Convert paid from integer to boolean
            bill['paid'] = bool(bill['paid'])
            bills.append(bill)
        
        conn.close()
        success_msg(f"Loaded {len(bills)} bills from database")
        
    except Exception as e:
        error_msg(f"Error loading bills from database: {e}")
        bills = []

def save_bills():
    """Save bills to SQLite database."""
    try:
        # Initialize database if it doesn't exist
        initialize_database()
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Clear existing bills
        cursor.execute('DELETE FROM bills')
        
        # Insert all bills
        for bill in bills:
            cursor.execute('''
                INSERT INTO bills (
                    name, due_date, billing_cycle, reminder_days, web_page,
                    login_info, password, paid, company_email, support_phone,
                    billing_phone, customer_service_hours, account_number,
                    reference_id, support_chat_url, mobile_app
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                bill.get('name', ''),
                bill.get('due_date', ''),
                bill.get('billing_cycle', 'monthly'),
                bill.get('reminder_days', 7),
                bill.get('web_page', ''),
                bill.get('login_info', ''),
                bill.get('password', ''),
                1 if bill.get('paid', False) else 0,
                bill.get('company_email', ''),
                bill.get('support_phone', ''),
                bill.get('billing_phone', ''),
                bill.get('customer_service_hours', ''),
                bill.get('account_number', ''),
                bill.get('reference_id', ''),
                bill.get('support_chat_url', ''),
                bill.get('mobile_app', '')
            ))
        
        conn.commit()
        conn.close()
        success_msg("Bills saved to database successfully")
        
    except Exception as e:
        error_msg(f"Save error: {e}")

def backup_bills():
    """Main backup function with progress."""
    backup_bills_with_progress()

def backup_bills_with_progress():
    """Create backup with progress indicator."""
    try:
        # Step 1: Check directories
        with ProgressBar.create_bar(100, "🔍 Checking directories", "blue") as pbar:
            if not os.path.exists(BACKUP_DIR):
                os.makedirs(BACKUP_DIR)
            pbar.update(50)
            
            if not os.path.exists(BILLS_FILE):
                info_msg("No bills.json found. No backup needed.")
                pbar.update(100)
                return
            pbar.update(100)
        
        # Step 2: Create backup
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_name = f"bills_backup_{timestamp}.json"
        backup_path = os.path.join(BACKUP_DIR, backup_name)
        
        with ProgressBar.create_bar(100, "💾 Creating backup", "green") as pbar:
            # Simulate file copy with progress
            file_size = os.path.getsize(BILLS_FILE)
            chunk_size = max(1024, file_size // 10)  # 10 chunks minimum
            
            with open(BILLS_FILE, 'rb') as src, open(backup_path, 'wb') as dst:
                copied = 0
                while True:
                    chunk = src.read(chunk_size)
                    if not chunk:
                        break
                    dst.write(chunk)
                    copied += len(chunk)
                    progress = min(95, (copied / file_size) * 100)
                    pbar.n = progress
                    pbar.refresh()
                    time.sleep(0.01)  # Small delay for visual effect
                pbar.update(100 - pbar.n)
        
        success_msg(f"Backup created: {backup_name}")
        
        # Step 3: Cleanup old backups
        cleanup_old_backups_with_progress()
        
    except Exception as e:
        error_msg(f"Backup error: {e}")

def cleanup_old_backups_with_progress():
    """Remove old backups with progress indicator."""
    try:
        with ProgressBar.create_bar(100, "🧹 Cleaning old backups", "yellow") as pbar:
            backup_files = [f for f in os.listdir(BACKUP_DIR) 
                           if f.startswith('bills_backup_')]
            pbar.update(30)
            
            backup_files.sort(key=lambda x: os.path.getmtime(
                os.path.join(BACKUP_DIR, x)))
            pbar.update(60)
            
            removed_count = 0
            while len(backup_files) > MAX_BACKUPS:
                oldest = backup_files.pop(0)
                os.remove(os.path.join(BACKUP_DIR, oldest))
                removed_count += 1
                info_msg(f"Removed old backup: {oldest}")
                time.sleep(0.1)  # Small delay for visual feedback
            
            pbar.update(100)
            
            if removed_count > 0:
                success_msg(f"Cleaned up {removed_count} old backup(s)")
            else:
                info_msg("No old backups to clean")
                
    except Exception as e:
        warning_msg(f"Cleanup error: {e}")

def save_bills_with_progress():
    """Save bills with progress indicator."""
    try:
        # Create backup first
        backup_bills_with_progress()
        
        # Save bills with progress
        with ProgressBar.create_bar(100, "💾 Saving bills", "green") as pbar:
            pbar.update(20)
            
            # Convert to JSON with progress simulation
            json_data = json.dumps(bills, indent=2)
            pbar.update(60)
            
            # Write to file
            with open(BILLS_FILE, 'w') as f:
                f.write(json_data)
            pbar.update(100)
            
        success_msg("Bills saved successfully")
        
    except Exception as e:
        error_msg(f"Save error: {e}")

# 6. Input validation functions
def get_required_input(prompt):
    """Get required input with cancel option and colors."""
    while True:
        value = colored_input(f"{prompt}: ", Colors.PROMPT).strip()
        if value.lower() == 'cancel':
            return None
        if value:
            return value
        error_msg("This field is required. Please enter a value or type 'cancel' to cancel.")

def get_optional_input(prompt):
    """Get optional input with cancel option and colors."""
    value = colored_input(f"{prompt} (optional): ", Colors.PROMPT).strip()
    if value.lower() == 'cancel':
        return None
    return value or ""

def get_valid_date(prompt):
    """Get a valid date input."""
    while True:
        date_str = get_required_input(prompt)
        if date_str is None:  # User cancelled
            return None
        try:
            datetime.strptime(date_str, DATE_FORMAT)
            return date_str
        except ValueError:
            print(f"❌ Invalid date format. Please use {DATE_FORMAT}")

def get_yes_no(prompt):
    """Get yes/no input."""
    while True:
        response = input(prompt).strip().lower()
        if response in ['yes', 'y']:
            return True
        elif response in ['no', 'n']:
            return False
        print("Please enter 'yes' or 'no'")

# 6.2 Enhanced validation functions
# Import the comprehensive validation module
from validation import DataValidator, ValidationError

def validate_url(url):
    """Legacy URL validation function for backward compatibility."""
    is_valid, error_msg, cleaned_url = DataValidator.validate_url(url)
    return cleaned_url if is_valid else None

def validate_email(email):
    """Legacy email validation function for backward compatibility."""
    is_valid, error_msg = DataValidator.validate_email(email)
    return email.strip().lower() if is_valid else None

def validate_date_range(start_date, end_date):
    """Validate that start_date is before end_date."""
    try:
        start = datetime.strptime(start_date, DATE_FORMAT)
        end = datetime.strptime(end_date, DATE_FORMAT)
        return start <= end
    except ValueError:
        return False

def validate_reminder_days(days_str):
    """Legacy reminder days validation function for backward compatibility."""
    is_valid, error_msg = DataValidator.validate_reminder_days(days_str)
    return is_valid

def validate_future_date(date_str):
    """Legacy future date validation function for backward compatibility."""
    return DataValidator.validate_due_date(date_str)

def get_valid_url(prompt):
    """Get a valid URL input with enhanced validation."""
    while True:
        url = colored_input(f"{prompt}: ", Colors.PROMPT).strip()
        
        if url.lower() == 'cancel':
            return None
            
        if not url:  # Empty URL is allowed
            return ""
            
        is_valid, error_msg_text, validated_url = DataValidator.validate_url(url)
        if is_valid:
            if validated_url != url:
                success_msg(f"URL corrected to: {validated_url}")
            return validated_url
        else:
            error_msg(f"Invalid URL: {error_msg_text}")

def get_valid_email(prompt):
    """Get a valid email input with enhanced validation."""
    while True:
        email = colored_input(f"{prompt}: ", Colors.PROMPT).strip()
        
        if email.lower() == 'cancel':
            return None
            
        if not email:  # Empty email is allowed
            return ""
            
        is_valid, error_msg_text = DataValidator.validate_email(email)
        if is_valid:
            return email.strip().lower()
        else:
            error_msg(f"Invalid email: {error_msg_text}")

def get_valid_reminder_days(prompt, default=7):
    """Get valid reminder days with enhanced validation."""
    while True:
        days_input = colored_input(f"{prompt} [default: {default}]: ", Colors.PROMPT).strip()
        
        if days_input.lower() == 'cancel':
            return None
            
        if not days_input:
            return default
            
        is_valid, error_msg_text = DataValidator.validate_reminder_days(days_input)
        if is_valid:
            return int(days_input)
        else:
            error_msg(f"Invalid reminder days: {error_msg_text}")

def get_valid_date_with_range_check(prompt):
    """Get a valid date input with enhanced range validation."""
    while True:
        date_str = get_required_input(prompt)
        if date_str is None:  # User cancelled
            return None
            
        is_valid, error_msg_text = DataValidator.validate_due_date(date_str)
        if is_valid:
            return date_str
        else:
            error_msg(f"Invalid date: {error_msg_text}")

# 6.1 Billing cycle constants and functions
class BillingCycle:
    """Billing cycle constants and utilities."""
    WEEKLY = "weekly"
    BI_WEEKLY = "bi-weekly"
    MONTHLY = "monthly"
    QUARTERLY = "quarterly"
    SEMI_ANNUALLY = "semi-annually"
    ANNUALLY = "annually"
    ONE_TIME = "one-time"
    
    @staticmethod
    def get_all_cycles():
        return [
            BillingCycle.WEEKLY,
            BillingCycle.BI_WEEKLY,
            BillingCycle.MONTHLY,
            BillingCycle.QUARTERLY,
            BillingCycle.SEMI_ANNUALLY,
            BillingCycle.ANNUALLY,
            BillingCycle.ONE_TIME
        ]
    
    @staticmethod
    def get_cycle_description(cycle):
        descriptions = {
            BillingCycle.WEEKLY: "Every 7 days",
            BillingCycle.BI_WEEKLY: "Every 14 days",
            BillingCycle.MONTHLY: "Every month",
            BillingCycle.QUARTERLY: "Every 3 months",
            BillingCycle.SEMI_ANNUALLY: "Every 6 months",
            BillingCycle.ANNUALLY: "Every 12 months",
            BillingCycle.ONE_TIME: "One-time payment (no recurrence)"
        }
        return descriptions.get(cycle, "Unknown cycle")

def get_billing_cycle():
    """Get billing cycle from user input."""
    print("\n--- Select Billing Cycle ---")
    cycles = BillingCycle.get_all_cycles()
    
    for idx, cycle in enumerate(cycles, 1):
        description = BillingCycle.get_cycle_description(cycle)
        print(f"{idx}. {cycle.title()} - {description}")
    
    while True:
        try:
            choice = colored_input(f"\nChoose billing cycle (1-{len(cycles)}) or 'cancel': ", Colors.PROMPT).strip()
            
            if choice.lower() == 'cancel':
                return None
            
            choice_num = int(choice)
            if 1 <= choice_num <= len(cycles):
                selected_cycle = cycles[choice_num - 1]
                success_msg(f"Selected: {selected_cycle.title()}")
                return selected_cycle
            else:
                error_msg(f"Please choose a number between 1 and {len(cycles)}")
        except ValueError:
            error_msg("Please enter a valid number or 'cancel'")

def calculate_next_due_date(current_due_date, billing_cycle):
    """Calculate the next due date based on billing cycle."""
    try:
        current_date = datetime.strptime(current_due_date, DATE_FORMAT)
    except ValueError:
        return current_due_date  # Return original if can't parse
    
    if billing_cycle == BillingCycle.WEEKLY:
        next_date = current_date + timedelta(days=7)
    elif billing_cycle == BillingCycle.BI_WEEKLY:
        next_date = current_date + timedelta(days=14)
    elif billing_cycle == BillingCycle.MONTHLY:
        # Handle month rollover properly
        next_date = add_months(current_date, 1)
    elif billing_cycle == BillingCycle.QUARTERLY:
        next_date = add_months(current_date, 3)
    elif billing_cycle == BillingCycle.SEMI_ANNUALLY:
        next_date = add_months(current_date, 6)
    elif billing_cycle == BillingCycle.ANNUALLY:
        next_date = add_months(current_date, 12)
    elif billing_cycle == BillingCycle.ONE_TIME:
        return current_due_date  # No change for one-time bills
    else:
        return current_due_date  # Unknown cycle, no change
    
    return next_date.strftime(DATE_FORMAT)

def add_months(date, months):
    """Add months to a date, handling month/year rollover properly."""
    month = date.month - 1 + months
    year = date.year + month // 12
    month = month % 12 + 1
    
    # Handle day overflow (e.g., Jan 31 + 1 month should be Feb 28/29)
    day = date.day
    max_day = calendar.monthrange(year, month)[1]
    if day > max_day:
        day = max_day
    
    return date.replace(year=year, month=month, day=day)

def get_billing_cycle_color(cycle):
    """Get color for billing cycle display."""
    color_map = {
        BillingCycle.WEEKLY: Colors.DUE_SOON,
        BillingCycle.BI_WEEKLY: Colors.WARNING,
        BillingCycle.MONTHLY: Colors.INFO,
        BillingCycle.QUARTERLY: Colors.SUCCESS,
        BillingCycle.SEMI_ANNUALLY: Colors.TITLE,
        BillingCycle.ANNUALLY: Colors.MENU,
        BillingCycle.ONE_TIME: Colors.ERROR
    }
    return color_map.get(cycle, Colors.RESET)

# 7. Core bill management functions
def add_bill():
    """Add a new bill with colored feedback and auto-complete assistance."""
    title_msg("Add a New Bill")
    info_msg("Type 'cancel' at any time to cancel.")
    
    # Show existing bills for reference if any exist
    if bills:
        print(f"\n{Colors.INFO}📋 Existing bills for reference:{Colors.RESET}")
        for i, bill in enumerate(bills[:5], 1):  # Show first 5 bills
            status = "✓ Paid" if bill.get('paid', False) else "○ Unpaid"
            print(f"{Colors.INFO}  • {bill['name']} [{status}]{Colors.RESET}")
        if len(bills) > 5:
            print(f"{Colors.INFO}  ... and {len(bills) - 5} more bills{Colors.RESET}")
        print()
    
    # Ask for bill name and check for duplicates
    while True:
        name = get_required_input("Enter the name of the new bill")
        if name is None:
            warning_msg("Bill addition cancelled.")
            return
        
        # Check for duplicates
        if any(bill['name'].lower() == name.lower() for bill in bills):
            error_msg(f"A bill with the name '{name}' already exists. Please enter a different name.")
            # Show similar names for reference
            similar_names = AutoComplete.suggest_names(name, max_suggestions=3)
            if similar_names:
                print(f"{Colors.WARNING}💡 Similar existing bills:{Colors.RESET}")
                for similar in similar_names:
                    print(f"{Colors.WARNING}  • {similar}{Colors.RESET}")
        else:
            break

    # Get due date with validation
    due_date = get_valid_date_with_range_check("Enter the due date of the bill (YYYY-MM-DD)")
    if due_date is None:
        warning_msg("Bill addition cancelled.")
        return
    
    # Get billing cycle
    billing_cycle = get_billing_cycle()
    if billing_cycle is None:
        warning_msg("Bill addition cancelled.")
        return
    
    # Get reminder period
    reminder_days = get_valid_reminder_days("Enter reminder days before due date (1-365)")
    if reminder_days is None:
        warning_msg("Bill addition cancelled.")
        return
    
    # Get optional fields with validation
    web_page = get_valid_url("Enter the web page for the bill (optional)")
    if web_page is None:
        warning_msg("Bill addition cancelled.")
        return
    
    login_info = get_optional_input("Enter the login information for the bill")
    if login_info is None:
        warning_msg("Bill addition cancelled.")
        return

    password = get_optional_input("Enter the password for the bill")
    if password is None:
        warning_msg("Bill addition cancelled.")
        return

    # Get contact information
    print(f"\n{Colors.TITLE}📞 Contact Information (Optional){Colors.RESET}")
    print(f"{Colors.INFO}Add customer service contact details for this bill:{Colors.RESET}")
    
    company_email = get_valid_email("Enter company customer service email")
    if company_email is None:
        warning_msg("Bill addition cancelled.")
        return
    
    support_phone = get_optional_input("Enter customer support phone number")
    if support_phone is None:
        warning_msg("Bill addition cancelled.")
        return
    
    billing_phone = get_optional_input("Enter billing department phone number")
    if billing_phone is None:
        warning_msg("Bill addition cancelled.")
        return
    
    customer_service_hours = get_optional_input("Enter customer service hours (e.g., Mon-Fri 9AM-5PM)")
    if customer_service_hours is None:
        warning_msg("Bill addition cancelled.")
        return
    
    account_number = get_optional_input("Enter account/customer number")
    if account_number is None:
        warning_msg("Bill addition cancelled.")
        return
    
    reference_id = get_optional_input("Enter reference/policy number")
    if reference_id is None:
        warning_msg("Bill addition cancelled.")
        return
    
    support_chat_url = get_valid_url("Enter live chat support URL (optional)")
    if support_chat_url is None:
        warning_msg("Bill addition cancelled.")
        return
    
    mobile_app = get_optional_input("Enter mobile app information (e.g., 'Netflix App - iOS/Android')")
    if mobile_app is None:
        warning_msg("Bill addition cancelled.")
        return

    # Create bill data
    bill_data = {
        "name": name,
        "due_date": due_date,
        "web_page": web_page,
        "login_info": login_info,
        "password": password,
        "paid": False,
        "billing_cycle": billing_cycle,
        "reminder_days": reminder_days,
        # Contact information
        "company_email": company_email,
        "support_phone": support_phone,
        "billing_phone": billing_phone,
        "customer_service_hours": customer_service_hours,
        "account_number": account_number,
        "reference_id": reference_id,
        "support_chat_url": support_chat_url,
        "mobile_app": mobile_app
    }
    
    # Validate complete bill data before saving
    is_valid, error_msg_text, cleaned_data = DataValidator.validate_bill_data(bill_data)
    if not is_valid:
        error_msg(f"Bill validation failed: {error_msg_text}")
        warning_msg("Bill addition cancelled due to validation errors.")
        return
    
    # Add validated bill to list
    bills.append(cleaned_data)
    save_bills()
    success_msg(f"Bill '{name}' added successfully with {billing_cycle} billing cycle!")
    colored_input("Press Enter to continue...", Colors.INFO)

def display_menu():
    """Display the main menu with colors."""
    print("\n" + Colors.MENU + "="*40)
    title_msg("BILLS TRACKER")
    print(Colors.MENU + "="*40 + Colors.RESET)
    
    print(f"{Colors.MENU}1.{Colors.RESET} 📝 Add a bill")
    print(f"{Colors.MENU}2.{Colors.RESET} 📋 View all bills")
    print(f"{Colors.MENU}3.{Colors.RESET} 🔍 Search bills")
    print(f"{Colors.MENU}4.{Colors.RESET} 🔄 Sort bills")
    print(f"{Colors.MENU}5.{Colors.RESET} ⏰ Check due bills")
    print(f"{Colors.MENU}6.{Colors.RESET} 💰 Pay a bill")
    print(f"{Colors.MENU}7.{Colors.RESET} ✏️  Edit a bill")
    print(f"{Colors.MENU}8.{Colors.RESET} 🗑️  Delete a bill")
    print(f"{Colors.MENU}9.{Colors.RESET} 📋 Bill templates")
    print(f"{Colors.MENU}10.{Colors.RESET} 📥 CSV Import/Export")
    print(f"{Colors.MENU}11.{Colors.RESET} 🔐 Password Management")
    print(f"{Colors.MENU}12.{Colors.RESET} 🔍 Data Integrity Check")
    print(f"{Colors.MENU}13.{Colors.RESET} 🗜️  Data Compression")
    print(f"{Colors.MENU}14.{Colors.RESET} 📖 Help")
    print(f"{Colors.MENU}15.{Colors.RESET} 🚪 Exit")
    print(Colors.MENU + "="*40 + Colors.RESET)

def view_bills():
    """View all bills with color coding."""
    title_msg("All Bills")
    
    if not bills:
        warning_msg("No bills found.")
        return
    
    today = datetime.now()
    
    for idx, bill in enumerate(bills, 1):
        # Determine bill status and color
        if bill.get('paid', False):
            status = f"{Colors.PAID}✓ Paid{Colors.RESET}"
        else:
            status = f"{Colors.UNPAID}○ Unpaid{Colors.RESET}"
        
        # Calculate days until due
        try:
            due_date = datetime.strptime(bill['due_date'], DATE_FORMAT)
            days_diff = (due_date - today).days
            
            if days_diff < 0:
                date_info = f"{Colors.OVERDUE}(Overdue by {abs(days_diff)} days!){Colors.RESET}"
                status = f"{Colors.OVERDUE}! OVERDUE{Colors.RESET}"
            elif days_diff == 0:
                date_info = f"{Colors.DUE_SOON}(Due TODAY!){Colors.RESET}"
            elif days_diff <= 7:
                date_info = f"{Colors.DUE_SOON}(Due in {days_diff} days){Colors.RESET}"
            else:
                date_info = ""
        except ValueError:
            date_info = f"{Colors.ERROR}(Invalid date){Colors.RESET}"
        
        # Print bill info with colors
        print(f"{Colors.INFO}{idx:2}.{Colors.RESET} {Colors.TITLE}{bill['name']}{Colors.RESET} [{status}]")
        print(f"    Due: {Colors.INFO}{bill['due_date']}{Colors.RESET} {date_info}")
        
        # Show billing cycle
        cycle = bill.get('billing_cycle', 'monthly')
        cycle_color = get_billing_cycle_color(cycle)
        print(f"    Cycle: {cycle_color}{cycle.title()}{Colors.RESET}")
        
        # Show reminder period
        reminder_days = bill.get('reminder_days', 7)
        if reminder_days == 1:
            reminder_text = "1 day before"
        else:
            reminder_text = f"{reminder_days} days before"
        print(f"    Reminder: {Colors.WARNING}⏰ {reminder_text}{Colors.RESET}")
        
        if bill.get('web_page'):
            print(f"    Website: {Colors.INFO}{bill['web_page']}{Colors.RESET}")
        if bill.get('login_info'):
            print(f"    Login: {Colors.INFO}{bill['login_info']}{Colors.RESET}")
        
        # Show contact information if available
        contact_info = []
        if bill.get('company_email'):
            contact_info.append(f"📧 {bill['company_email']}")
        if bill.get('support_phone'):
            contact_info.append(f"📞 Support: {bill['support_phone']}")
        if bill.get('billing_phone'):
            contact_info.append(f"💰 Billing: {bill['billing_phone']}")
        if bill.get('account_number'):
            contact_info.append(f"🆔 Account: {bill['account_number']}")
        
        if contact_info:
            print(f"    {Colors.INFO}📞 Contact: {', '.join(contact_info[:2])}{'...' if len(contact_info) > 2 else ''}{Colors.RESET}")
        
        print()

def edit_bill():
    print("\n--- Edit a Bill ---")
    view_bills()
    if not bills:
        return
    try:
        choice = int(input("Enter the number of the bill to edit:"))
        if 1 <= choice <= len(bills):
            bill = bills[choice - 1]
            print(f"Editing '{bill['name']}'")
            new_name = input(f"Name [{bill['name']}]: ").strip()
            if new_name:
                bill['name'] = new_name

            new_due_date = input(f"Due Date [{bill['due_date']}]: ").strip()
            if new_due_date:
                # Validate date format and range
                try:
                    datetime.strptime(new_due_date, '%Y-%m-%d')
                    is_valid, error_msg_text = validate_future_date(new_due_date)
                    if is_valid:
                        bill['due_date'] = new_due_date
                    else:
                        error_msg(error_msg_text + " Keeping the original date.")
                except ValueError:
                    error_msg("Invalid date format. Keeping the original date.")
                
            # Website with validation
            print(f"Current website: {bill.get('web_page', 'Not provided')}")
            new_web_page = input(f"Web Page [{bill['web_page']}]: ").strip()
            if new_web_page:
                if new_web_page.lower() == 'clear':
                    bill['web_page'] = ""
                    success_msg("Website cleared.")
                else:
                    validated_url = validate_url(new_web_page)
                    if validated_url is not None:
                        bill['web_page'] = validated_url
                        if validated_url != new_web_page:
                            success_msg(f"Website corrected to: {validated_url}")
                    else:
                        error_msg("Invalid URL format. Keeping the original website.")

            new_login_info = input(f"Login Info [{bill['login_info']}]: ").strip()
            if new_login_info:
                bill['login_info'] = new_login_info

            # Show decrypted password for editing
            current_password = bill.get('password', '')
            if current_password and CRYPTOGRAPHY_AVAILABLE and password_encryption.fernet:
                current_password = password_encryption.decrypt_password(current_password)
            
            new_password = input(f"Password [{current_password}]: ").strip()
            if new_password:
                bill['password'] = new_password

            paid_status = input(f"Paid (yes/no) [{ 'yes' if bill.get('paid', False) else 'no' }]: ").strip().lower()
            if paid_status in ['yes', 'y']:
                bill['paid'] = True
            elif paid_status in ['no', 'n']:
                bill['paid'] = False

            # Billing cycle
            current_cycle = bill.get('billing_cycle', 'monthly')
            print(f"\nCurrent billing cycle: {current_cycle.title()}")
            change_cycle = input("Change billing cycle? (yes/no): ").strip().lower()
            if change_cycle in ['yes', 'y']:
                new_billing_cycle = get_billing_cycle()
                if new_billing_cycle is not None:
                    bill['billing_cycle'] = new_billing_cycle
                    success_msg(f"Billing cycle updated to {new_billing_cycle}")

            # Reminder period
            current_reminder = bill.get('reminder_days', 7)
            print(f"\nCurrent reminder period: {current_reminder} days before due date")
            change_reminder = input("Change reminder period? (yes/no): ").strip().lower()
            if change_reminder in ['yes', 'y']:
                new_reminder_days = get_valid_reminder_days("Enter new reminder days before due date (1-365)", bill.get('reminder_days', 7))
                if new_reminder_days is not None:
                    bill['reminder_days'] = new_reminder_days
                    success_msg(f"Reminder period updated to {new_reminder_days} days")

            # Contact Information
            print(f"\n{Colors.TITLE}📞 Contact Information{Colors.RESET}")
            
            new_company_email = colored_input(f"Company Email [{bill.get('company_email', '')}]: ", Colors.PROMPT).strip()
            if new_company_email:
                if new_company_email.lower() == 'clear':
                    bill['company_email'] = ""
                    success_msg("Company email cleared.")
                else:
                    validated_email = validate_email(new_company_email)
                    if validated_email is not None:
                        bill['company_email'] = validated_email
                    else:
                        error_msg("Invalid email format. Keeping the original email.")

            new_support_phone = colored_input(f"Support Phone [{bill.get('support_phone', '')}]: ", Colors.PROMPT).strip()
            if new_support_phone:
                bill['support_phone'] = new_support_phone

            new_billing_phone = colored_input(f"Billing Phone [{bill.get('billing_phone', '')}]: ", Colors.PROMPT).strip()
            if new_billing_phone:
                bill['billing_phone'] = new_billing_phone

            new_service_hours = colored_input(f"Service Hours [{bill.get('customer_service_hours', '')}]: ", Colors.PROMPT).strip()
            if new_service_hours:
                bill['customer_service_hours'] = new_service_hours

            new_account_number = colored_input(f"Account Number [{bill.get('account_number', '')}]: ", Colors.PROMPT).strip()
            if new_account_number:
                bill['account_number'] = new_account_number

            new_reference_id = colored_input(f"Reference ID [{bill.get('reference_id', '')}]: ", Colors.PROMPT).strip()
            if new_reference_id:
                bill['reference_id'] = new_reference_id

            new_support_chat_url = colored_input(f"Support Chat URL [{bill.get('support_chat_url', '')}]: ", Colors.PROMPT).strip()
            if new_support_chat_url:
                if new_support_chat_url.lower() == 'clear':
                    bill['support_chat_url'] = ""
                    success_msg("Support chat URL cleared.")
                else:
                    validated_url = validate_url(new_support_chat_url)
                    if validated_url is not None:
                        bill['support_chat_url'] = validated_url
                    else:
                        error_msg("Invalid URL format. Keeping the original URL.")

            new_mobile_app = colored_input(f"Mobile App [{bill.get('mobile_app', '')}]: ", Colors.PROMPT).strip()
            if new_mobile_app:
                bill['mobile_app'] = new_mobile_app

            save_bills()
            success_msg(f"Bill '{bill['name']}' updated successfully.")
        else:
            print("Invalid selection.")
    except ValueError:
        print("Invalid input. Please enter a number.")

def delete_bill():
    print("\n--- Delete a Bill ---")
    view_bills()
    if not bills:
        return
    try:
        choice = int(input("Enter the number of the bill to delete:"))
        if 1 <= choice <= len(bills):
            bill = bills[choice - 1]
            confirm = input(f"Are you sure you want to delete '{bill['name']}'? (yes/no): ").strip().lower()
            if confirm in ['yes', 'y']:
                bills.remove(bill)
                save_bills()
                print(f"Bill '{bill['name']}' has been deleted.")
            else:
                print("Deletion cancelled.")
        else:
            print("Invalid selection.")
    except ValueError:
        print("Invalid input.")
        print("Invalid input. Please enter a number.")

def pay_bill():
    """Pay a bill with enhanced user experience."""
    print("\n--- Pay a Bill ---")
    view_bills()
    if not bills:
        return
    
    # Filter unpaid bills for better user experience
    unpaid_bills = [bill for bill in bills if not bill.get('paid', False)]
    if not unpaid_bills:
        warning_msg("All bills are already paid! 🎉")
        input("Press Enter to continue...")
        return
    
    try:
        choice = int(colored_input("Enter the number of the bill you want to pay: ", Colors.PROMPT))
        if 1 <= choice <= len(bills):
            bill = bills[choice - 1]
            if bill.get("paid", False):
                warning_msg(f"Bill '{bill['name']}' has already been paid.")
                input("Press Enter to continue...")
                return
            
            # Get billing cycle info
            billing_cycle = bill.get('billing_cycle', BillingCycle.MONTHLY)
            
            if billing_cycle == BillingCycle.ONE_TIME:
                # One-time bills just get marked as paid
                bill["paid"] = True
                success_msg(f"One-time bill '{bill['name']}' marked as paid and completed!")
            else:
                # For recurring bills, ask user preference
                print(f"\n'{bill['name']}' is a {billing_cycle} recurring bill.")
                print("Payment options:")
                print("1. 💰 Pay and advance to next billing cycle (recommended)")
                print("2. ✅ Mark as permanently paid (stops recurring)")
                
                payment_choice = colored_input("Choose option (1-2): ", Colors.PROMPT).strip()
                
                if payment_choice == '1':
                    # Standard recurring bill payment
                    old_due_date = bill['due_date']
                    new_due_date = calculate_next_due_date(bill['due_date'], billing_cycle)
                    bill['due_date'] = new_due_date
                    bill["paid"] = False  # Unpaid for next cycle
                    
                    success_msg(f"Bill '{bill['name']}' marked as paid!")
                    info_msg(f"Next due date ({billing_cycle}): {new_due_date}")
                elif payment_choice == '2':
                    # Mark as permanently paid
                    bill["paid"] = True
                    success_msg(f"Bill '{bill['name']}' marked as permanently paid!")
                    warning_msg("This bill will no longer appear in due bills until you manually edit it.")
                else:
                    error_msg("Invalid choice. Payment cancelled.")
                    input("Press Enter to continue...")
                    return
            
            save_bills()
            colored_input("\nPress Enter to continue...", Colors.INFO)
        else:
            error_msg("Invalid selection. Please choose a valid bill number.")
            input("Press Enter to continue...")
    except ValueError:
        error_msg("Invalid input. Please enter a valid number.")
        input("Press Enter to continue...")

def verify_due_bills(days=None):
    """Check for due bills with color highlighting. If days=None, use each bill's custom reminder period."""
    if days is None:
        title_msg("Bills Due Based on Custom Reminder Periods")
    else:
        title_msg(f"Bills Due Within {days} Days")
    
    today = datetime.now()
    due_bills = []
    
    for bill in bills:
        if bill.get('paid', False):
            continue  # Skip paid bills
            
        try:
            due_date = datetime.strptime(bill['due_date'], DATE_FORMAT)
            days_diff = (due_date - today).days
            
            # Use custom reminder period if days parameter is None
            if days is None:
                reminder_period = bill.get('reminder_days', 7)
                if days_diff <= reminder_period:
                    due_bills.append((bill, days_diff, reminder_period))
            else:
                # Use provided days parameter
                if days_diff <= days:
                    due_bills.append((bill, days_diff, days))
        except ValueError:
            continue
    
    if not due_bills:
        if days is None:
            success_msg("No bills are due based on your custom reminder periods! 🎉")
        else:
            success_msg("No bills are due soon! 🎉")
        return
    
    # Sort by days until due
    due_bills.sort(key=lambda x: x[1])
    
    for bill, days_diff, reminder_period in due_bills:
        if days_diff < 0:
            colored_print(f"🚨 {bill['name']} - OVERDUE by {abs(days_diff)} days!", Colors.OVERDUE)
        elif days_diff == 0:
            colored_print(f"🔥 {bill['name']} - DUE TODAY!", Colors.DUE_SOON)
        elif days_diff <= 3:
            colored_print(f"⚠️  {bill['name']} - Due in {days_diff} days", Colors.WARNING)
        else:
            colored_print(f"📅 {bill['name']} - Due in {days_diff} days", Colors.INFO)
        
        # Show reminder period for custom reminders
        if days is None:
            colored_print(f"   ⏰ Reminder set for {reminder_period} days before due date", Colors.INFO)
    print()

def verify_due_bills_paginated(days=None, items_per_page=10):
    """Check for due bills with pagination. If days=None, use each bill's custom reminder period."""
    if days is None:
        title_msg("Bills Due Based on Custom Reminder Periods")
    else:
        title_msg(f"Bills Due Within {days} Days")
    
    if not bills:
        warning_msg("No bills found.")
        return
    
    # Get due bills
    due_bills = get_due_bills(days)
    
    if not due_bills:
        if days is None:
            success_msg("No bills are due based on your custom reminder periods! 🎉")
        else:
            success_msg("No bills are due soon! 🎉")
        input("Press Enter to continue...")
        return
    
    # Sort by urgency (overdue first, then by days)
    due_bills.sort(key=lambda x: (x[1] >= 0, x[1]))  # Overdue first, then by days
    
    paginator = Paginator(due_bills, items_per_page)
    
    while True:
        clear_console()
        title_msg(f"Bills Due Within {days} Days")
        
        # Display current page of due bills
        current_due_bills = paginator.get_page()
        display_due_bills_page(current_due_bills, paginator)
        
        # Display pagination controls
        display_pagination_controls(paginator)
        
        # Additional options
        print("\nDue Bills Options:")
        print("1. 💰 Pay a bill from this list")
        print("2. 💳 Pay multiple bills")
        print("3. 📅 Change days filter")
        
        choice = colored_input("\nEnter your choice: ", Colors.PROMPT).strip().lower()
        
        if choice == 'n' and paginator.get_page_info()['has_next']:
            paginator.next_page()
        elif choice == 'p' and paginator.get_page_info()['has_prev']:
            paginator.prev_page()
        elif choice == 'g':
            goto_page(paginator)
        elif choice == 's':
            new_size = change_page_size()
            if new_size:
                paginator = Paginator(due_bills, new_size)
        elif choice == '1':
            pay_bill_from_due_list_paginated(current_due_bills, paginator)
        elif choice == '2':
            bulk_pay_bills_from_due_list(due_bills)
        elif choice == '3':
            new_days = change_days_filter()
            if new_days:
                return verify_due_bills_paginated(new_days, items_per_page)
        elif choice == 'b':
            break
        else:
            error_msg("Invalid option.")
            input("Press Enter to continue...")

def get_due_bills(days=None):
    """Get bills due within specified days. If days=None, use each bill's custom reminder period."""
    today = datetime.now()
    due_bills = []
    
    for bill in bills:
        if bill.get('paid', False):
            continue  # Skip paid bills
            
        try:
            due_date = datetime.strptime(bill['due_date'], DATE_FORMAT)
            days_diff = (due_date - today).days
            
            # Use custom reminder period if days parameter is None
            if days is None:
                reminder_period = bill.get('reminder_days', 7)
                if days_diff <= reminder_period:
                    due_bills.append((bill, days_diff))
            else:
                # Use provided days parameter
                if days_diff <= days:
                    due_bills.append((bill, days_diff))
        except ValueError:
            continue  # Skip invalid dates
    
    return due_bills

def display_due_bills_page(current_due_bills, paginator):
    """Display a page of due bills with urgency indicators."""
    page_info = paginator.get_page_info()
    
    success_msg(f"Found {page_info['total_items']} bills due soon:")
    print()
    
    for idx, (bill, days_diff) in enumerate(current_due_bills, 1):
        actual_number = (paginator.current_page - 1) * paginator.items_per_page + idx
        
        # Determine bill status and color
        if bill.get('paid', False):
            status = f"{Colors.PAID}✓ Paid{Colors.RESET}"
        else:
            status = f"{Colors.UNPAID}○ Unpaid{Colors.RESET}"
        
        # Override status for overdue bills
        if days_diff < 0:
            status = f"{Colors.OVERDUE}! OVERDUE{Colors.RESET}"
        
        # Determine urgency and color
        if days_diff < 0:
            urgency = f"{Colors.OVERDUE}🚨 OVERDUE by {abs(days_diff)} days!{Colors.RESET}"
        elif days_diff == 0:
            urgency = f"{Colors.DUE_SOON}🔥 DUE TODAY!{Colors.RESET}"
        elif days_diff <= 3:
            urgency = f"{Colors.WARNING}⚠️  Due in {days_diff} days{Colors.RESET}"
        else:
            urgency = f"{Colors.INFO}📅 Due in {days_diff} days{Colors.RESET}"
        
        print(f"{Colors.INFO}{actual_number:3}.{Colors.RESET} {Colors.TITLE}{bill['name']}{Colors.RESET} [{status}]")
        print(f"     {urgency}")
        print(f"     Due Date: {Colors.INFO}{bill['due_date']}{Colors.RESET}")
        if bill.get('web_page'):
            print(f"     Website: {Colors.INFO}{bill['web_page'][:40]}{'...' if len(bill.get('web_page', '')) > 40 else ''}{Colors.RESET}")
        if bill.get('login_info'):
            print(f"     Login: {Colors.INFO}{bill['login_info'][:30]}{'...' if len(bill.get('login_info', '')) > 30 else ''}{Colors.RESET}")
        print()

def change_days_filter():
    """Change the number of days for due bill filter."""
    try:
        days = int(colored_input("Enter number of days to check (1-365): ", Colors.PROMPT))
        if 1 <= days <= 365:
            success_msg(f"Filter changed to {days} days")
            return days
        else:
            error_msg("Days must be between 1 and 365")
    except ValueError:
        error_msg("Please enter a valid number")
    
    input("Press Enter to continue...")
    return None

# 8. Search functions
def search_bills():
    """Search bills by name, due date, or website."""
    print("\n--- Search Bills ---")
    print("Search options:")
    print("1. Search by name")
    print("2. Search by due date")
    print("3. Search by website")
    print("4. Search by contact information")
    print("5. Search all fields")
    print("6. Back to main menu")
    
    choice = input("\nChoose search option (1-6): ").strip()
    
    if choice == '1':
        search_by_name()
    elif choice == '2':
        search_by_due_date()
    elif choice == '3':
        search_by_website()
    elif choice == '4':
        search_by_contact_info()
    elif choice == '5':
        search_all_fields()
    elif choice == '6':
        return
    else:
        print("❌ Invalid option. Please choose 1-6.")
        input("Press Enter to continue...")
        search_bills()

def search_by_name():
    """Search bills by name with auto-complete suggestions."""
    if not bills:
        warning_msg("No bills found.")
        return
    
    colored_print("\n🔍 Search Bills by Name", Colors.TITLE)
    
    search_term = get_input_with_autocomplete(
        "Enter bill name to search (with auto-complete)", 
        autocomplete_type="bills", 
        allow_empty=False
    )
    
    if search_term is None:
        warning_msg("Search cancelled.")
        return
    
    if not search_term:
        error_msg("Search term cannot be empty.")
        return
    
    # Perform search (case-insensitive partial match)
    search_term_lower = search_term.lower()
    results = [bill for bill in bills 
              if search_term_lower in bill['name'].lower()]
    
    display_search_results(results, f"Bills containing '{search_term}' in name")

def search_by_due_date():
    """Search bills by due date or date range."""
    if not bills:
        print("No bills found.")
        return
    
    print("\nSearch options:")
    print("1. Exact date (YYYY-MM-DD)")
    print("2. Month and year (YYYY-MM)")
    print("3. Year only (YYYY)")
    
    option = input("Choose option (1-3): ").strip()
    
    if option == '1':
        search_term = input("Enter exact date (YYYY-MM-DD): ").strip()
        results = [bill for bill in bills if bill['due_date'] == search_term]
        display_search_results(results, f"Bills due on {search_term}")
        
    elif option == '2':
        search_term = input("Enter month and year (YYYY-MM): ").strip()
        results = [bill for bill in bills if bill['due_date'].startswith(search_term)]
        display_search_results(results, f"Bills due in {search_term}")
        
    elif option == '3':
        search_term = input("Enter year (YYYY): ").strip()
        results = [bill for bill in bills if bill['due_date'].startswith(search_term)]
        display_search_results(results, f"Bills due in {search_term}")
    else:
        print("❌ Invalid option.")

def search_by_website():
    """Search bills by website with auto-complete suggestions."""
    if not bills:
        warning_msg("No bills found.")
        return
    
    websites = AutoComplete.get_websites()
    if not websites:
        warning_msg("No websites found in bills.")
        return
    
    colored_print("\n🌐 Search Bills by Website", Colors.TITLE)
    
    search_term = get_input_with_autocomplete(
        "Enter website to search (with auto-complete)", 
        autocomplete_type="websites", 
        allow_empty=False
    )
    
    if search_term is None:
        warning_msg("Search cancelled.")
        return
    
    if not search_term:
        error_msg("Search term cannot be empty.")
        return
    
    # Perform search (case-insensitive partial match)
    search_term_lower = search_term.lower()
    results = [bill for bill in bills 
              if search_term_lower in bill.get('web_page', '').lower()]
    
    display_search_results(results, f"Bills with website containing '{search_term}'")

def search_by_contact_info():
    """Search bills by contact information."""
    if not bills:
        warning_msg("No bills found.")
        return
    
    print("\nSearch by Contact Information:")
    print("1. Search by company email")
    print("2. Search by phone number")
    print("3. Search by account number")
    print("4. Search by reference ID")
    print("5. Back to search menu")
    
    choice = input("\nChoose option (1-5): ").strip()
    
    if choice == '1':
        search_term = input("Enter company email to search: ").strip().lower()
        results = [bill for bill in bills 
                  if search_term in bill.get('company_email', '').lower()]
        display_search_results(results, f"Bills with company email containing '{search_term}'")
        
    elif choice == '2':
        search_term = input("Enter phone number to search: ").strip()
        results = [bill for bill in bills 
                  if (search_term in bill.get('support_phone', '') or 
                      search_term in bill.get('billing_phone', ''))]
        display_search_results(results, f"Bills with phone number containing '{search_term}'")
        
    elif choice == '3':
        search_term = input("Enter account number to search: ").strip()
        results = [bill for bill in bills 
                  if search_term in bill.get('account_number', '')]
        display_search_results(results, f"Bills with account number containing '{search_term}'")
        
    elif choice == '4':
        search_term = input("Enter reference ID to search: ").strip()
        results = [bill for bill in bills 
                  if search_term in bill.get('reference_id', '')]
        display_search_results(results, f"Bills with reference ID containing '{search_term}'")
        
    elif choice == '5':
        return
    else:
        error_msg("Invalid option. Please choose 1-5.")
        input("Press Enter to continue...")
        search_by_contact_info()

def search_all_fields_with_progress(search_term):
    """Search across all bill fields with progress."""
    if not bills:
        warning_msg("No bills found.")
        return []
    
    results = []
    search_term_lower = search_term.lower()
    
    with ProgressBar.create_bar(len(bills), "🔍 Searching bills", "blue") as pbar:
        for bill in bills:
            # Search in all text fields
            searchable_text = ' '.join([
                bill.get('name', ''),
                bill.get('due_date', ''),
                bill.get('web_page', ''),
                bill.get('login_info', ''),
                bill.get('company_email', ''),
                bill.get('support_phone', ''),
                bill.get('billing_phone', ''),
                bill.get('customer_service_hours', ''),
                bill.get('account_number', ''),
                bill.get('reference_id', ''),
                bill.get('support_chat_url', ''),
                bill.get('mobile_app', '')
            ]).lower()
            
            if search_term_lower in searchable_text:
                results.append(bill)
            
            pbar.update(1)
            time.sleep(0.02)  # Small delay for visual effect
    
    return results

def search_all_fields():
    """Search across all bill fields with progress indicator."""
    if not bills:
        warning_msg("No bills found.")
        return
    
    search_term = input("Enter search term (searches all fields): ").strip()
    if not search_term:
        error_msg("Search term cannot be empty.")
        return
    
    results = search_all_fields_with_progress(search_term)
    display_search_results(results, f"Bills containing '{search_term}' in any field")

def display_search_results(results, title):
    """Display search results with automatic pagination."""
    if len(results) > 10:
        display_search_results_paginated(results, title)
    else:
        display_search_results_simple(results, title)

def display_search_results_simple(results, title):
    """Display search results without pagination (for small result sets)."""
    colored_print(f"\n--- {title} ---", Colors.TITLE)
    
    if not results:
        error_msg("No bills found matching your search.")
        input("\nPress Enter to continue...")
        return
    
    success_msg(f"Found {len(results)} bill(s):")
    print()
    
    today = datetime.now()
    
    for idx, bill in enumerate(results, 1):
        # Determine bill status and color (same logic as view_bills)
        if bill.get('paid', False):
            status = f"{Colors.PAID}✓ Paid{Colors.RESET}"
        else:
            status = f"{Colors.UNPAID}○ Unpaid{Colors.RESET}"
        
        # Calculate days until due
        try:
            due_date = datetime.strptime(bill['due_date'], DATE_FORMAT)
            days_diff = (due_date - today).days
            
            if days_diff < 0:
                date_info = f"{Colors.OVERDUE}(Overdue by {abs(days_diff)} days!){Colors.RESET}"
                status = f"{Colors.OVERDUE}! OVERDUE{Colors.RESET}"
            elif days_diff == 0:
                date_info = f"{Colors.DUE_SOON}(Due TODAY!){Colors.RESET}"
            elif days_diff <= 7:
                date_info = f"{Colors.DUE_SOON}(Due in {days_diff} days){Colors.RESET}"
            else:
                date_info = ""
        except ValueError:
            date_info = f"{Colors.ERROR}(Invalid date){Colors.RESET}"
        
        # Display bill with colors (same format as view_bills)
        print(f"{Colors.INFO}{idx:2}.{Colors.RESET} {Colors.TITLE}{bill['name']}{Colors.RESET} [{status}]")
        print(f"    Due: {Colors.INFO}{bill['due_date']}{Colors.RESET} {date_info}")
        
        if bill.get('web_page'):
            print(f"    Website: {Colors.INFO}{bill['web_page']}{Colors.RESET}")
        if bill.get('login_info'):
            print(f"    Login: {Colors.INFO}{bill['login_info']}{Colors.RESET}")
        print()
    
    # Keep the existing options for simple display
    print("Options:")
    print("1. View details of a specific bill")
    print("2. Pay a bill from results")
    print("3. Back to search menu")
    
    choice = input("Choose option (1-3): ").strip()
    
    if choice == '1':
        view_bill_details_from_search(results)
    elif choice == '2':
        pay_bill_from_search(results)
    elif choice == '3':
        return
    else:
        error_msg("Invalid option.")
        input("Press Enter to continue...")

def display_search_results_paginated(results, title, items_per_page=10):
    """Display search results with pagination."""
    if not results:
        print(f"\n--- {title} ---")
        error_msg("No bills found matching your search.")
        input("\nPress Enter to continue...")
        return
    
    paginator = Paginator(results, items_per_page)
    
    while True:
        clear_console()
        colored_print(f"--- {title} ---", Colors.TITLE)
        
        # Display current page of results
        current_results = paginator.get_page()
        display_search_results_page(current_results, paginator)
        
        # Display pagination controls
        display_pagination_controls(paginator)
        
        # Additional options for search results
        print("\nSearch Result Options:")
        print("1. 👁️  View details of a bill")
        print("2. 💰 Pay a bill from results")
        print("3. ✏️  Edit a bill from results")
        
        # Get user input
        choice = colored_input("\nEnter your choice: ", Colors.PROMPT).strip().lower()
        
        if choice == 'n' and paginator.get_page_info()['has_next']:
            paginator.next_page()
        elif choice == 'p' and paginator.get_page_info()['has_prev']:
            paginator.prev_page()
        elif choice == 'g':
            goto_page(paginator)
        elif choice == 's':
            new_size = change_page_size()
            if new_size:
                paginator = Paginator(results, new_size)
        elif choice == '1':
            view_bill_details_from_search_paginated(current_results, paginator)
        elif choice == '2':
            pay_bill_from_search_paginated(current_results, paginator)
        elif choice == '3':
            edit_bill_from_search_paginated(current_results, paginator)
        elif choice == 'b':
            break
        else:
            error_msg("Invalid option.")
            input("Press Enter to continue...")

def display_search_results_page(current_results, paginator):
    """Display a page of search results with proper color coding."""
    page_info = paginator.get_page_info()
    
    success_msg(f"Found {page_info['total_items']} bill(s) total:")
    print()
    
    today = datetime.now()
    
    for idx, bill in enumerate(current_results, 1):
        # Calculate actual bill number across all pages
        actual_number = (paginator.current_page - 1) * paginator.items_per_page + idx
        
        # Determine bill status and color (same logic as view_bills)
        if bill.get('paid', False):
            status = f"{Colors.PAID}✓ Paid{Colors.RESET}"
        else:
            status = f"{Colors.UNPAID}○ Unpaid{Colors.RESET}"
        
        # Calculate days until due
        try:
            due_date = datetime.strptime(bill['due_date'], DATE_FORMAT)
            days_diff = (due_date - today).days
            
            if days_diff < 0:
                date_info = f"{Colors.OVERDUE}(Overdue by {abs(days_diff)} days!){Colors.RESET}"
                status = f"{Colors.OVERDUE}! OVERDUE{Colors.RESET}"
            elif days_diff == 0:
                date_info = f"{Colors.DUE_SOON}(Due TODAY!){Colors.RESET}"
            elif days_diff <= 7:
                date_info = f"{Colors.DUE_SOON}(Due in {days_diff} days){Colors.RESET}"
            else:
                date_info = ""
        except ValueError:
            date_info = f"{Colors.ERROR}(Invalid date){Colors.RESET}"
        
        # Display bill with colors (same format as view_bills)
        print(f"{Colors.INFO}{actual_number:3}.{Colors.RESET} {Colors.TITLE}{bill['name']}{Colors.RESET} [{status}]")
        print(f"     Due: {Colors.INFO}{bill['due_date']}{Colors.RESET} {date_info}")
        
        if bill.get('web_page'):
            print(f"     Website: {Colors.INFO}{bill['web_page'][:50]}{'...' if len(bill.get('web_page', '')) > 50 else ''}{Colors.RESET}")
        if bill.get('login_info'):
            print(f"     Login: {Colors.INFO}{bill['login_info'][:30]}{'...' if len(bill.get('login_info', '')) > 30 else ''}{Colors.RESET}")
        print()

# Add these after your color utility functions

class ProgressBar:
    """Progress bar utility functions."""
    
    @staticmethod
    def create_bar(total, description="Processing", color="green"):
        """Create a progress bar with custom styling."""
        bar_format = "{desc}: {percentage:3.0f}%|{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}]"
        
        # Color mapping for tqdm
        color_map = {
            "green": "\033[92m",
            "blue": "\033[94m", 
            "yellow": "\033[93m",
            "red": "\033[91m",
            "cyan": "\033[96m"
        }
        
        return tqdm(
            total=total,
            desc=f"{color_map.get(color, '')}{description}\033[0m",
            bar_format=bar_format,
            ncols=70,
            ascii=True,
            colour=color
        )
    
    @staticmethod
    def simulate_work(pbar, steps, work_function=None):
        """Simulate work with progress updates."""
        step_size = 100 / steps
        for i in range(steps):
            if work_function:
                work_function()
            else:
                time.sleep(0.1)  # Simulate work
            pbar.update(step_size)

def show_progress(func, description="Processing", steps=10, color="green"):
    """Decorator to show progress bar for functions."""
    def wrapper(*args, **kwargs):
        with ProgressBar.create_bar(100, description, color) as pbar:
            # Simulate progress during function execution
            result = func(*args, **kwargs)
            pbar.update(100)
            return result
    return wrapper

# 9. Pagination utility classes and functions
class Paginator:
    """Pagination utility for handling large datasets."""
    
    def __init__(self, items, items_per_page=10):
        self.items = items
        self.items_per_page = items_per_page
        self.total_items = len(items)
        self.total_pages = max(1, (self.total_items + items_per_page - 1) // items_per_page)
        self.current_page = 1
    
    def get_page(self, page_number=None):
        """Get items for a specific page."""
        if page_number is not None:
            self.current_page = max(1, min(page_number, self.total_pages))
        
        start_idx = (self.current_page - 1) * self.items_per_page
        end_idx = start_idx + self.items_per_page
        return self.items[start_idx:end_idx]
    
    def next_page(self):
        """Go to next page."""
        if self.current_page < self.total_pages:
            self.current_page += 1
            return True
        return False
    
    def prev_page(self):
        """Go to previous page."""
        if self.current_page > 1:
            self.current_page -= 1
            return True
        return False
    
    def get_page_info(self):
        """Get pagination information."""
        start_item = (self.current_page - 1) * self.items_per_page + 1
        end_item = min(self.current_page * self.items_per_page, self.total_items)
        
        return {
            'current_page': self.current_page,
            'total_pages': self.total_pages,
            'total_items': self.total_items,
            'start_item': start_item,
            'end_item': end_item,
            'has_next': self.current_page < self.total_pages,
            'has_prev': self.current_page > 1
        }

def display_pagination_controls(paginator):
    """Display pagination controls and information."""
    info = paginator.get_page_info()
    
    # Display page information
    colored_print(
        f"📄 Page {info['current_page']} of {info['total_pages']} "
        f"(showing {info['start_item']}-{info['end_item']} of {info['total_items']} bills)",
        Colors.INFO
    )
    
    # Display navigation options
    print("\n" + "="*50)
    nav_options = []
    
    if info['has_prev']:
        nav_options.append("⬅️  [P] Previous page")
    if info['has_next']:
        nav_options.append("➡️  [N] Next page")
    
    nav_options.extend([
        "🔢 [G] Go to page",
        "📊 [S] Change page size",
        "🚪 [B] Back to main menu"
    ])
    
    for option in nav_options:
        print(option)
    print("="*50)

def due_bills_menu():
    """Display due bills menu options."""
    title_msg("Due Bills Options")
    print(f"{Colors.MENU}1.{Colors.RESET} ⏰ Check bills with custom reminder periods")
    print(f"{Colors.MENU}2.{Colors.RESET} 📅 Check bills due within specific days")
    print(f"{Colors.MENU}3.{Colors.RESET} 🔙 Back to main menu")
    print()
    
    while True:
        choice = colored_input("Choose an option (1-3): ", Colors.PROMPT).strip()
        
        if choice == '1':
            clear_console()
            # Use custom reminder periods
            if len(bills) > 10:
                verify_due_bills_paginated(days=None)
            else:
                verify_due_bills(days=None)
                colored_input("\n📖 Press Enter to continue...", Colors.INFO)
            break
        elif choice == '2':
            clear_console()
            # Ask for specific number of days
            try:
                custom_days = int(colored_input("Enter number of days to check (1-365): ", Colors.PROMPT))
                if 1 <= custom_days <= 365:
                    if len(bills) > 10:
                        verify_due_bills_paginated(days=custom_days)
                    else:
                        verify_due_bills(days=custom_days)
                        colored_input("\n📖 Press Enter to continue...", Colors.INFO)
                else:
                    error_msg("Please enter a number between 1 and 365")
                    continue
            except ValueError:
                error_msg("Please enter a valid number")
                continue
            break
        elif choice == '3':
            break
        else:
            error_msg("Please choose option 1, 2, or 3")

# 11. Pagination functions (moved here to be available before main)
def view_bills_paginated(items_per_page=10):
    """View all bills with pagination."""
    if not bills:
        warning_msg("No bills found.")
        input("Press Enter to continue...")
        return
    
    paginator = Paginator(bills, items_per_page)
    
    while True:
        clear_console()
        title_msg("All Bills")
        
        # Display current page of bills
        current_bills = paginator.get_page()
        display_bills_page(current_bills, paginator)
        
        # Display pagination controls
        display_pagination_controls(paginator)
        
        # Get user input
        choice = colored_input("\nEnter your choice: ", Colors.PROMPT).strip().lower()
        
        if choice == 'n' and paginator.get_page_info()['has_next']:
            paginator.next_page()
        elif choice == 'p' and paginator.get_page_info()['has_prev']:
            paginator.prev_page()
        elif choice == 'g':
            goto_page(paginator)
        elif choice == 's':
            new_size = change_page_size()
            if new_size:
                paginator = Paginator(bills, new_size)
        elif choice == 'b':
            break
        else:
            error_msg("Invalid option. Use P/N to navigate, G to go to page, S to change size, B to go back.")
            input("Press Enter to continue...")

def display_bills_page(current_bills, paginator):
    """Display a page of bills with enhanced formatting."""
    today = datetime.now()
    page_info = paginator.get_page_info()
    
    for idx, bill in enumerate(current_bills, page_info['start_item']):
        # Determine bill status and color
        if bill.get('paid', False):
            status = f"{Colors.PAID}✓ Paid{Colors.RESET}"
        else:
            status = f"{Colors.UNPAID}○ Unpaid{Colors.RESET}"
        
        # Calculate days until due
        try:
            due_date = datetime.strptime(bill['due_date'], DATE_FORMAT)
            days_diff = (due_date - today).days
            
            if days_diff < 0:
                date_info = f"{Colors.OVERDUE}(Overdue by {abs(days_diff)} days!){Colors.RESET}"
                status = f"{Colors.OVERDUE}! OVERDUE{Colors.RESET}"
            elif days_diff == 0:
                date_info = f"{Colors.DUE_SOON}(Due TODAY!){Colors.RESET}"
            elif days_diff <= 7:
                date_info = f"{Colors.DUE_SOON}(Due in {days_diff} days){Colors.RESET}"
            else:
                date_info = ""
        except ValueError:
            date_info = f"{Colors.ERROR}(Invalid date){Colors.RESET}"
        
        # Print bill info with colors and numbers
        print(f"{Colors.INFO}{idx:3}.{Colors.RESET} {Colors.TITLE}{bill['name']}{Colors.RESET} [{status}]")
        print(f"     Due: {Colors.INFO}{bill['due_date']}{Colors.RESET} {date_info}")
        
        if bill.get('web_page'):
            print(f"     Website: {Colors.INFO}{bill['web_page'][:50]}{'...' if len(bill.get('web_page', '')) > 50 else ''}{Colors.RESET}")
        if bill.get('login_info'):
            print(f"     Login: {Colors.INFO}{bill['login_info'][:30]}{'...' if len(bill.get('login_info', '')) > 30 else ''}{Colors.RESET}")
        print()

def goto_page(paginator):
    """Go to a specific page."""
    info = paginator.get_page_info()
    try:
        page = int(colored_input(f"Enter page number (1-{info['total_pages']}): ", Colors.PROMPT))
        if 1 <= page <= info['total_pages']:
            paginator.get_page(page)
            success_msg(f"Jumped to page {page}")
        else:
            error_msg(f"Page must be between 1 and {info['total_pages']}")
    except ValueError:
        error_msg("Please enter a valid number")
    
    input("Press Enter to continue...")

def change_page_size():
    """Change the number of items per page."""
    try:
        size = int(colored_input("Enter items per page (5-50): ", Colors.PROMPT))
        if 5 <= size <= 50:
            success_msg(f"Page size changed to {size}")
            return size
        else:
            error_msg("Page size must be between 5 and 50")
    except ValueError:
        error_msg("Please enter a valid number")
    
    input("Press Enter to continue...")
    return None

def display_bill_details(bill):
    """Display detailed information for a single bill."""
    print(f"\n--- Bill Details: {bill['name']} ---")
    print(f"Name: {Colors.TITLE}{bill['name']}{Colors.RESET}")
    print(f"Due Date: {Colors.INFO}{bill['due_date']}{Colors.RESET}")
    
    status_color = Colors.PAID if bill.get('paid', False) else Colors.UNPAID
    status_text = "✓ Paid" if bill.get('paid', False) else "○ Unpaid"
    print(f"Status: {status_color}{status_text}{Colors.RESET}")
    
    print(f"Website: {Colors.INFO}{bill.get('web_page', 'Not provided')}{Colors.RESET}")
    print(f"Login Info: {Colors.INFO}{bill.get('login_info', 'Not provided')}{Colors.RESET}")
    
    # Show password as asterisks (decrypt if encrypted)
    password = bill.get('password', '')
    if password and CRYPTOGRAPHY_AVAILABLE and password_encryption.fernet:
        # Decrypt password for display
        decrypted_password = password_encryption.decrypt_password(password)
        password_display = '*' * len(decrypted_password) if decrypted_password else 'Not provided'
    else:
        password_display = '*' * len(password) if password else 'Not provided'
        print(f"Password: {Colors.INFO}{password_display}{Colors.RESET}")
    
    # Show billing cycle and reminder
    cycle = bill.get('billing_cycle', 'monthly')
    cycle_color = get_billing_cycle_color(cycle)
    print(f"Billing Cycle: {cycle_color}{cycle.title()}{Colors.RESET}")
    
    reminder_days = bill.get('reminder_days', 7)
    print(f"Reminder: {Colors.WARNING}{reminder_days} days before due date{Colors.RESET}")
    
    # Show contact information
    print(f"\n{Colors.TITLE}📞 Contact Information:{Colors.RESET}")
    if bill.get('company_email'):
        print(f"  📧 Customer Service Email: {Colors.INFO}{bill['company_email']}{Colors.RESET}")
    if bill.get('support_phone'):
        print(f"  📞 Support Phone: {Colors.INFO}{bill['support_phone']}{Colors.RESET}")
    if bill.get('billing_phone'):
        print(f"  💰 Billing Phone: {Colors.INFO}{bill['billing_phone']}{Colors.RESET}")
    if bill.get('customer_service_hours'):
        print(f"  🕒 Service Hours: {Colors.INFO}{bill['customer_service_hours']}{Colors.RESET}")
    if bill.get('account_number'):
        print(f"  🆔 Account Number: {Colors.INFO}{bill['account_number']}{Colors.RESET}")
    if bill.get('reference_id'):
        print(f"  📋 Reference ID: {Colors.INFO}{bill['reference_id']}{Colors.RESET}")
    if bill.get('support_chat_url'):
        print(f"  💬 Live Chat: {Colors.INFO}{bill['support_chat_url']}{Colors.RESET}")
    if bill.get('mobile_app'):
        print(f"  📱 Mobile App: {Colors.INFO}{bill['mobile_app']}{Colors.RESET}")
    
    # Check if no contact info was provided
    contact_fields = ['company_email', 'support_phone', 'billing_phone', 'customer_service_hours', 
                     'account_number', 'reference_id', 'support_chat_url', 'mobile_app']
    if not any(bill.get(field) for field in contact_fields):
        print(f"  {Colors.WARNING}No contact information provided{Colors.RESET}")

def run_data_integrity_check():
    """Run a comprehensive data integrity check and provide repair options."""
    clear_console()
    title_msg("🔍 Data Integrity Check")
    print("=" * 50)
    
    info_msg("Running comprehensive data integrity check...")
    print()
    
    # Create integrity checker
    integrity_checker = DataIntegrityChecker(DB_FILE)
    
    # Run the check
    is_healthy, issues = integrity_checker.check_database_integrity()
    
    # Display the report
    integrity_checker.print_report()
    
    if not is_healthy:
        print("\n" + "=" * 50)
        print("🔧 Repair Options:")
        print("1. Attempt automatic repairs")
        print("2. View detailed issues")
        print("3. Return to main menu")
        
        repair_choice = colored_input("\nChoose an option (1-3): ", Colors.PROMPT).strip()
        
        if repair_choice == '1':
            info_msg("🔧 Attempting automatic repairs...")
            repairs = integrity_checker.repair_issues(auto_repair=True)
            
            if repairs:
                success_msg("✅ Repairs completed successfully!")
                print("\nRepairs made:")
                for i, repair in enumerate(repairs, 1):
                    print(f"  {i}. {repair}")
                
                # Run integrity check again to confirm
                info_msg("\n🔍 Running integrity check again...")
                is_healthy, issues = integrity_checker.check_database_integrity()
                
                if is_healthy:
                    success_msg("✅ Database is now healthy!")
                else:
                    warning_msg("⚠️  Some issues remain after repair.")
                    integrity_checker.print_report()
            else:
                warning_msg("⚠️  No automatic repairs were possible.")
        
        elif repair_choice == '2':
            print("\n" + "=" * 50)
            print("📋 Detailed Issues:")
            for i, issue in enumerate(issues, 1):
                print(f"{i}. {issue}")
        
        # Always wait for user input before returning
        colored_input("\nPress Enter to return to main menu...", Colors.INFO)
    else:
        success_msg("✅ Database integrity check passed successfully!")
        colored_input("\nPress Enter to return to main menu...", Colors.INFO)

# 12. Main application
def main():
    """Main application loop with pagination support and master password protection."""
    clear_console()
    title_msg("Welcome to Bills Tracker! 🏠💳")

    # Require master password before proceeding
    master_password = verify_master_password()

    # Initialize encryption with master password
    if CRYPTOGRAPHY_AVAILABLE:
        password_encryption.initialize_encryption(master_password)

    # Start session timer
    start_session()
    
    # Run data integrity check on startup
    info_msg("🔍 Running data integrity check...")
    integrity_checker = DataIntegrityChecker(DB_FILE)
    is_healthy, issues = integrity_checker.check_database_integrity()
    
    if not is_healthy:
        warning_msg("⚠️  Data integrity issues found!")
        integrity_checker.print_report()
        
        # Ask user if they want to attempt repairs
        repair_choice = get_yes_no("Would you like to attempt automatic repairs? (y/n): ")
        if repair_choice:
            info_msg("🔧 Attempting automatic repairs...")
            repairs = integrity_checker.repair_issues(auto_repair=True)
            
            if repairs:
                success_msg("✅ Repairs completed successfully!")
                # Run integrity check again to confirm
                is_healthy, issues = integrity_checker.check_database_integrity()
                if is_healthy:
                    success_msg("✅ Database is now healthy!")
                else:
                    warning_msg("⚠️  Some issues remain after repair.")
            else:
                warning_msg("⚠️  No automatic repairs were possible.")
        
        colored_input("Press Enter to continue...", Colors.INFO)
    else:
        success_msg("✅ Data integrity check passed!")
    
    # Load existing bills and templates
    load_bills()
    load_templates()

    while True:
        display_menu()
        choice = colored_input("Choose an option (1-15): ", Colors.PROMPT).strip()
        
        if choice == '1':
            clear_console()
            add_bill()
        elif choice == '2':
            clear_console()
            # Use pagination if more than 10 bills
            if len(bills) > 10:
                view_bills_paginated()
            else:
                view_bills()
                colored_input("\n📖 Press Enter to continue...", Colors.INFO)
        elif choice == '3':
            clear_console()
            search_bills()
        elif choice == '4':
            clear_console()
            sort_bills()
        elif choice == '5':
            clear_console()
            due_bills_menu()
        elif choice == '6':
            clear_console()
            pay_bill()
        elif choice == '7':
            clear_console()
            edit_bill()
        elif choice == '8':
            clear_console()
            delete_bill()
        elif choice == '9':
            clear_console()
            templates_menu()
        elif choice == '10':
            clear_console()
            csv_import_export_menu()
        elif choice == '11':
            clear_console()
            password_management_menu()
        elif choice == '12':
            clear_console()
            run_data_integrity_check()
        elif choice == '13':
            clear_console()
            data_compression_menu()
        elif choice == '14':
            clear_console()
            show_help_menu()
        elif choice == '15':
            success_msg("Thank you for using Bills Tracker! 👋")
            break
        else:
            error_msg("Invalid option. Please choose 1-15.")
            colored_input("Press Enter to continue...", Colors.WARNING)

def data_compression_menu():
    """Display data compression menu."""
    compressor = DataCompressor()
    
    while True:
        clear_console()
        title_msg("🗜️ Data Compression")
        print("=" * 50)
        
        print(f"{Colors.MENU}1.{Colors.RESET} 📊 Compress database")
        print(f"{Colors.MENU}2.{Colors.RESET} 📁 Compress backup directory")
        print(f"{Colors.MENU}3.{Colors.RESET} 📄 Compress individual files")
        print(f"{Colors.MENU}4.{Colors.RESET} 📈 Analyze compression effectiveness")
        print(f"{Colors.MENU}5.{Colors.RESET} 📋 View compression information")
        print(f"{Colors.MENU}6.{Colors.RESET} 🔄 Decompress files")
        print(f"{Colors.MENU}7.{Colors.RESET} 🚪 Back to main menu")
        
        choice = colored_input("\nChoose option (1-7): ", Colors.PROMPT).strip()
        
        if choice == '1':
            clear_console()
            compress_database_option(compressor)
        elif choice == '2':
            clear_console()
            compress_backup_directory_option(compressor)
        elif choice == '3':
            clear_console()
            compress_individual_files_option(compressor)
        elif choice == '4':
            clear_console()
            analyze_compression_option(compressor)
        elif choice == '5':
            clear_console()
            view_compression_info_option(compressor)
        elif choice == '6':
            clear_console()
            decompress_files_option(compressor)
        elif choice == '7':
            break
        else:
            error_msg("Invalid option. Please choose 1-7.")
            colored_input("Press Enter to continue...", Colors.WARNING)

def compress_database_option(compressor):
    """Compress the SQLite database."""
    title_msg("📊 Compress Database")
    print("=" * 40)
    
    if not os.path.exists(DB_FILE):
        error_msg(f"Database file '{DB_FILE}' not found.")
        colored_input("Press Enter to continue...", Colors.INFO)
        return
    
    # Show current database size
    db_size = os.path.getsize(DB_FILE)
    info_msg(f"Current database size: {db_size:,} bytes ({db_size / 1024:.1f} KB)")
    
    # Choose compression method
    print(f"\n{Colors.TITLE}Compression Methods:{Colors.RESET}")
    print("1. GZIP - Fast compression, good ratio")
    print("2. LZMA - Slower compression, best ratio")
    print("3. ZLIB - Balanced compression")
    print("4. Analyze best method")
    print("5. Back to compression menu")
    
    choice = colored_input("\nChoose compression method (1-5): ", Colors.PROMPT).strip()
    
    if choice == '1':
        method = 'gzip'
    elif choice == '2':
        method = 'lzma'
    elif choice == '3':
        method = 'zlib'
    elif choice == '4':
        # Analyze best method
        info_msg("Analyzing compression effectiveness...")
        analysis = compressor.analyze_compression_effectiveness(DB_FILE)
        
        if 'error' in analysis:
            error_msg(f"Analysis failed: {analysis['error']}")
            colored_input("Press Enter to continue...", Colors.INFO)
            return
        
        print(f"\n{Colors.TITLE}Compression Analysis Results:{Colors.RESET}")
        print(f"Original size: {analysis['original_size']:,} bytes")
        print()
        
        for method_name, data in analysis['methods'].items():
            if 'error' not in data:
                print(f"{method_name.upper()}:")
                print(f"  Compressed size: {data['compressed_size']:,} bytes")
                print(f"  Compression ratio: {data['compression_ratio']:.1f}%")
                print(f"  Compression time: {data['compression_time']:.2f} seconds")
                print()
        
        if analysis.get('best_method'):
            info_msg(f"Best method: {analysis['best_method'].upper()}")
            method = analysis['best_method']
        else:
            error_msg("No compression method available.")
            colored_input("Press Enter to continue...", Colors.INFO)
            return
    elif choice == '5':
        return
    else:
        error_msg("Invalid option.")
        colored_input("Press Enter to continue...", Colors.INFO)
        return
    
    # Confirm compression
    warning_msg(f"⚠️  This will compress the database using {method.upper()} compression.")
    confirm = colored_input("Continue? (yes/no): ", Colors.WARNING).strip().lower()
    
    if confirm not in ['yes', 'y']:
        info_msg("Compression cancelled.")
        colored_input("Press Enter to continue...", Colors.INFO)
        return
    
    # Compress database
    info_msg(f"Compressing database with {method.upper()}...")
    success, compressed_path, stats = compressor.compress_database(DB_FILE, method)
    
    if success:
        success_msg(f"Database compressed successfully!")
        print(f"\n{Colors.TITLE}Compression Results:{Colors.RESET}")
        print(f"Original size: {stats['original_size']:,} bytes")
        print(f"Compressed size: {stats['compressed_size']:,} bytes")
        print(f"Compression ratio: {stats['compression_ratio']:.1f}%")
        print(f"Space saved: {stats['original_size'] - stats['compressed_size']:,} bytes")
        print(f"Compression time: {stats['compression_time']:.2f} seconds")
        print(f"Compressed file: {compressed_path}")
        
        if stats.get('backup_created'):
            info_msg("Original database backed up before compression.")
    else:
        error_msg(f"Compression failed: {stats.get('error', 'Unknown error')}")
    
    colored_input("Press Enter to continue...", Colors.INFO)

def compress_backup_directory_option(compressor):
    """Compress all files in backup directory."""
    title_msg("📁 Compress Backup Directory")
    print("=" * 40)
    
    if not os.path.exists(BACKUP_DIR):
        error_msg(f"Backup directory '{BACKUP_DIR}' not found.")
        colored_input("Press Enter to continue...", Colors.INFO)
        return
    
    # Show backup directory info
    backup_files = [f for f in os.listdir(BACKUP_DIR) if os.path.isfile(os.path.join(BACKUP_DIR, f))]
    total_size = sum(os.path.getsize(os.path.join(BACKUP_DIR, f)) for f in backup_files)
    
    info_msg(f"Backup directory contains {len(backup_files)} files")
    info_msg(f"Total size: {total_size:,} bytes ({total_size / 1024:.1f} KB)")
    
    # Choose compression method
    print(f"\n{Colors.TITLE}Compression Method:{Colors.RESET}")
    print("1. GZIP - Fast compression, good ratio")
    print("2. LZMA - Slower compression, best ratio")
    print("3. ZLIB - Balanced compression")
    
    choice = colored_input("\nChoose compression method (1-3): ", Colors.PROMPT).strip()
    
    if choice == '1':
        method = 'gzip'
    elif choice == '2':
        method = 'lzma'
    elif choice == '3':
        method = 'zlib'
    else:
        error_msg("Invalid option.")
        colored_input("Press Enter to continue...", Colors.INFO)
        return
    
    # Ask about deleting originals
    print(f"\n{Colors.TITLE}Options:{Colors.RESET}")
    print("1. Keep original files (recommended)")
    print("2. Delete original files after compression")
    
    delete_choice = colored_input("\nChoose option (1-2): ", Colors.PROMPT).strip()
    delete_originals = delete_choice == '2'
    
    if delete_originals:
        warning_msg("⚠️  Original files will be deleted after compression!")
        confirm = colored_input("Continue? (yes/no): ", Colors.WARNING).strip().lower()
        if confirm not in ['yes', 'y']:
            info_msg("Compression cancelled.")
            colored_input("Press Enter to continue...", Colors.INFO)
            return
    
    # Compress backup directory
    info_msg(f"Compressing backup directory with {method.upper()}...")
    results = compressor.compress_backup_directory(BACKUP_DIR, method, delete_originals)
    
    if 'error' in results:
        error_msg(f"Compression failed: {results['error']}")
    else:
        success_msg("Backup directory compression completed!")
        print(f"\n{Colors.TITLE}Compression Results:{Colors.RESET}")
        print(f"Files processed: {results['files_processed']}")
        print(f"Files compressed: {results['files_compressed']}")
        print(f"Original size: {results['total_original_size']:,} bytes")
        print(f"Compressed size: {results['total_compressed_size']:,} bytes")
        print(f"Compression ratio: {results['compression_ratio']:.1f}%")
        print(f"Space saved: {results['total_original_size'] - results['total_compressed_size']:,} bytes")
        
        if results['errors']:
            print(f"\n{Colors.WARNING}Errors encountered:{Colors.RESET}")
            for error in results['errors']:
                print(f"  • {error}")
    
    colored_input("Press Enter to continue...", Colors.INFO)

def compress_individual_files_option(compressor):
    """Compress individual files."""
    title_msg("📄 Compress Individual Files")
    print("=" * 40)
    
    print("Enter file paths to compress (one per line, empty line to finish):")
    print("Example: bills_tracker.db")
    print("Example: backups/bills_backup_20241201_120000.json")
    print()
    
    file_paths = []
    while True:
        file_path = colored_input("File path (or empty to finish): ", Colors.PROMPT).strip()
        if not file_path:
            break
        if os.path.exists(file_path):
            file_paths.append(file_path)
        else:
            error_msg(f"File not found: {file_path}")
    
    if not file_paths:
        info_msg("No files to compress.")
        colored_input("Press Enter to continue...", Colors.INFO)
        return
    
    # Choose compression method
    print(f"\n{Colors.TITLE}Compression Method:{Colors.RESET}")
    print("1. GZIP - Fast compression, good ratio")
    print("2. LZMA - Slower compression, best ratio")
    print("3. ZLIB - Balanced compression")
    
    choice = colored_input("\nChoose compression method (1-3): ", Colors.PROMPT).strip()
    
    if choice == '1':
        method = 'gzip'
    elif choice == '2':
        method = 'lzma'
    elif choice == '3':
        method = 'zlib'
    else:
        error_msg("Invalid option.")
        colored_input("Press Enter to continue...", Colors.INFO)
        return
    
    # Ask about deleting originals
    delete_originals = colored_input("Delete original files after compression? (yes/no): ", Colors.WARNING).strip().lower() in ['yes', 'y']
    
    if delete_originals:
        warning_msg("⚠️  Original files will be deleted after compression!")
    
    # Compress files
    info_msg(f"Compressing {len(file_paths)} files with {method.upper()}...")
    results = compressor.batch_compress(file_paths, method, delete_originals)
    
    success_msg("Batch compression completed!")
    print(f"\n{Colors.TITLE}Compression Results:{Colors.RESET}")
    print(f"Total files: {results['total_files']}")
    print(f"Successful: {results['successful']}")
    print(f"Failed: {results['failed']}")
    print(f"Original size: {results['total_original_size']:,} bytes")
    print(f"Compressed size: {results['total_compressed_size']:,} bytes")
    print(f"Overall compression ratio: {results['overall_compression_ratio']:.1f}%")
    
    if results['errors']:
        print(f"\n{Colors.WARNING}Errors encountered:{Colors.RESET}")
        for error in results['errors']:
            print(f"  • {error}")
    
    colored_input("Press Enter to continue...", Colors.INFO)

def analyze_compression_option(compressor):
    """Analyze compression effectiveness for files."""
    title_msg("📈 Analyze Compression Effectiveness")
    print("=" * 40)
    
    file_path = colored_input("Enter file path to analyze: ", Colors.PROMPT).strip()
    
    if not os.path.exists(file_path):
        error_msg(f"File not found: {file_path}")
        colored_input("Press Enter to continue...", Colors.INFO)
        return
    
    info_msg("Analyzing compression effectiveness...")
    analysis = compressor.analyze_compression_effectiveness(file_path)
    
    if 'error' in analysis:
        error_msg(f"Analysis failed: {analysis['error']}")
        colored_input("Press Enter to continue...", Colors.INFO)
        return
    
    print(f"\n{Colors.TITLE}Compression Analysis Results:{Colors.RESET}")
    print(f"File: {file_path}")
    print(f"Original size: {analysis['original_size']:,} bytes ({analysis['original_size'] / 1024:.1f} KB)")
    print()
    
    for method_name, data in analysis['methods'].items():
        if 'error' not in data:
            print(f"{method_name.upper()}:")
            print(f"  Compressed size: {data['compressed_size']:,} bytes ({data['compressed_size'] / 1024:.1f} KB)")
            print(f"  Compression ratio: {data['compression_ratio']:.1f}%")
            print(f"  Space saved: {analysis['original_size'] - data['compressed_size']:,} bytes")
            print(f"  Compression time: {data['compression_time']:.2f} seconds")
            print()
        else:
            print(f"{method_name.upper()}: {data['error']}")
    
    if analysis.get('best_method'):
        success_msg(f"Best compression method: {analysis['best_method'].upper()}")
        success_msg(f"Best compression ratio: {analysis['best_compression_ratio']:.1f}%")
    
    colored_input("Press Enter to continue...", Colors.INFO)

def view_compression_info_option(compressor):
    """View compression information for files."""
    title_msg("📋 View Compression Information")
    print("=" * 40)
    
    file_path = colored_input("Enter file path: ", Colors.PROMPT).strip()
    
    if not os.path.exists(file_path):
        error_msg(f"File not found: {file_path}")
        colored_input("Press Enter to continue...", Colors.INFO)
        return
    
    info = compressor.get_compression_info(file_path)
    
    if 'error' in info:
        error_msg(f"Error: {info['error']}")
        colored_input("Press Enter to continue...", Colors.INFO)
        return
    
    print(f"\n{Colors.TITLE}File Information:{Colors.RESET}")
    print(f"File: {info['file_path']}")
    print(f"Size: {info['file_size']:,} bytes ({info['file_size'] / 1024:.1f} KB)")
    print(f"Is compressed: {'Yes' if info['is_compressed'] else 'No'}")
    
    if info['is_compressed']:
        print(f"Compression method: {info['compression_method'].upper()}")
        if info['original_size']:
            print(f"Original size: {info['original_size']:,} bytes")
            print(f"Compression ratio: {info['compression_ratio']:.1f}%")
            print(f"Space saved: {info['original_size'] - info['file_size']:,} bytes")
    
    colored_input("Press Enter to continue...", Colors.INFO)

def decompress_files_option(compressor):
    """Decompress compressed files."""
    title_msg("🔄 Decompress Files")
    print("=" * 40)
    
    file_path = colored_input("Enter compressed file path: ", Colors.PROMPT).strip()
    
    if not os.path.exists(file_path):
        error_msg(f"File not found: {file_path}")
        colored_input("Press Enter to continue...", Colors.INFO)
        return
    
    # Check if file is compressed
    info = compressor.get_compression_info(file_path)
    if not info['is_compressed']:
        error_msg("File is not compressed.")
        colored_input("Press Enter to continue...", Colors.INFO)
        return
    
    # Ask for output path
    output_path = colored_input("Output path (or empty for default): ", Colors.PROMPT).strip()
    if not output_path:
        output_path = None
    
    # Decompress file
    info_msg(f"Decompressing {file_path}...")
    success, decompressed_path = compressor.decompress_file(file_path, output_path)
    
    if success:
        success_msg(f"File decompressed successfully!")
        print(f"Decompressed file: {decompressed_path}")
        
        # Show size comparison
        original_size = os.path.getsize(file_path)
        decompressed_size = os.path.getsize(decompressed_path)
        print(f"Compressed size: {original_size:,} bytes")
        print(f"Decompressed size: {decompressed_size:,} bytes")
    else:
        error_msg(f"Decompression failed: {decompressed_path}")
    
    colored_input("Press Enter to continue...", Colors.INFO)

def show_help_menu():
    """Display comprehensive help information for the Bills Tracker application."""
    clear_console()
    title_msg("📚 Bills Tracker Help")
    print("=" * 60)
    
    print(f"\n{Colors.TITLE}🔧 Main Features:{Colors.RESET}")
    print("• Add, view, edit, and delete bills")
    print("• Track due dates with automatic notifications")
    print("• Flexible billing cycles (weekly, monthly, quarterly, etc.)")
    print("• Custom reminder periods per bill")
    print("• Password encryption and master password protection")
    print("• Automatic backup system")
    print("• Search and filter bills")
    print("• Import/export CSV and Excel files")
    print("• Bill templates for quick adding")
    print("• Contact information storage")
    print("• Data integrity checks")
    print("• Data compression for large datasets")
    
    print(f"\n{Colors.TITLE}📋 Menu Options:{Colors.RESET}")
    print("1. Add Bill - Create a new bill with all details")
    print("2. View Bills - Display all bills with pagination")
    print("3. Search Bills - Find bills by various criteria")
    print("4. Sort Bills - Arrange bills by different criteria")
    print("5. Due Bills - View bills due within specified days")
    print("6. Pay Bill - Mark bills as paid and update due dates")
    print("7. Edit Bill - Modify existing bill information")
    print("8. Delete Bill - Remove a bill from tracking")
    print("9. Templates - Save and reuse bill configurations")
    print("10. Import/Export - CSV and Excel file operations")
    print("11. Password Management - Secure password handling")
    print("12. Data Integrity - Check and repair data consistency")
    print("13. Data Compression - Compress large datasets")
    print("14. Help - This help menu")
    print("15. Exit - Close the application")
    
    print(f"\n{Colors.TITLE}🔐 Security Features:{Colors.RESET}")
    print("• Master password protection for all bill passwords")
    print("• Automatic session timeout after inactivity")
    print("• Encrypted password storage using Fernet encryption")
    print("• Secure password recovery options")
    print("• Data integrity verification on startup")
    
    print(f"\n{Colors.TITLE}📊 Data Management:{Colors.RESET}")
    print("• SQLite database for reliable data storage")
    print("• Automatic backup system with progress tracking")
    print("• CSV and Excel import/export with validation")
    print("• Data integrity checks and automatic repairs")
    print("• Bill templates for efficient data entry")
    print("• Data compression for storage optimization")
    
    print(f"\n{Colors.TITLE}🎯 Tips for Best Use:{Colors.RESET}")
    print("• Set up a master password on first use")
    print("• Use bill templates for recurring bills")
    print("• Set appropriate reminder periods for each bill")
    print("• Regularly check due bills to avoid late payments")
    print("• Use search and sort features to organize bills")
    print("• Export data regularly for backup purposes")
    print("• Run integrity checks if you notice data issues")
    print("• Use compression for large datasets to save space")
    
    print(f"\n{Colors.TITLE}🔧 Keyboard Shortcuts:{Colors.RESET}")
    print("• Use Tab for autocomplete suggestions")
    print("• Press Enter to accept suggestions")
    print("• Use arrow keys for navigation in paginated views")
    print("• Type 'q' to quit paginated views")
    print("• Use 'clear' to remove website/contact information")
    
    print(f"\n{Colors.TITLE}📞 Support:{Colors.RESET}")
    print("• Check the documentation in the 'docs' folder")
    print("• Review test files for usage examples")
    print("• Use data integrity checks for troubleshooting")
    print("• Export data before making major changes")
    
    print("\n" + "=" * 60)
    colored_input("\nPress Enter to return to main menu...", Colors.INFO)

# 10. Missing pagination helper functions
def view_bill_details_from_search(results):
    """View detailed information of a bill from search results."""
    try:
        choice = int(input(f"Enter bill number (1-{len(results)}): "))
        if 1 <= choice <= len(results):
            bill = results[choice - 1]
            display_bill_details(bill)
        else:
            error_msg("Invalid bill number.")
    except ValueError:
        error_msg("Please enter a valid number.")
    
    input("\nPress Enter to continue...")

def pay_bill_from_search(results):
    """Pay a bill from search results."""
    try:
        choice = int(input(f"Enter bill number to pay (1-{len(results)}): "))
        if 1 <= choice <= len(results):
            bill = results[choice - 1]
            if bill.get('paid', False):
                error_msg(f"Bill '{bill['name']}' is already paid.")
            else:
                # Find the bill in the main bills list and pay it
                for main_bill in bills:
                    if main_bill['name'] == bill['name'] and main_bill['due_date'] == bill['due_date']:
                        main_bill['paid'] = True
                        save_bills()
                        success_msg(f"Bill '{bill['name']}' marked as paid!")
                        break
        else:
            error_msg("Invalid bill number.")
    except ValueError:
        error_msg("Please enter a valid number.")
    
    input("Press Enter to continue...")

def view_bill_details_from_search_paginated(current_results, paginator):
    """View bill details from paginated search results."""
    if not current_results:
        return
    
    try:
        choice = int(colored_input(f"Enter bill number (1-{len(current_results)}): ", Colors.PROMPT))
        if 1 <= choice <= len(current_results):
            bill = current_results[choice - 1]
            display_bill_details(bill)
        else:
            error_msg("Invalid bill number.")
    except ValueError:
        error_msg("Please enter a valid number.")
    
    input("\nPress Enter to continue...")

def pay_bill_from_search_paginated(current_results, paginator):
    """Pay a bill from paginated search results."""
    if not current_results:
        return
    
    try:
        choice = int(colored_input(f"Enter bill number to pay (1-{len(current_results)}): ", Colors.PROMPT))
        if 1 <= choice <= len(current_results):
            bill = current_results[choice - 1]
            if bill.get('paid', False):
                error_msg(f"Bill '{bill['name']}' is already paid.")
            else:
                # Find the bill in the main bills list and pay it
                for main_bill in bills:
                    if main_bill['name'] == bill['name'] and main_bill['due_date'] == bill['due_date']:
                        main_bill['paid'] = True
                        save_bills()
                        success_msg(f"Bill '{bill['name']}' marked as paid!")
                        break
        else:
            error_msg("Invalid bill number.")
    except ValueError:
        error_msg("Please enter a valid number.")
    
    input("Press Enter to continue...")

def edit_bill_from_search_paginated(current_results, paginator):
    """Edit a bill from paginated search results."""
    if not current_results:
        return
    
    try:
        choice = int(colored_input(f"Enter bill number to edit (1-{len(current_results)}): ", Colors.PROMPT))
        if 1 <= choice <= len(current_results):
            bill = current_results[choice - 1]
            # Find the bill in the main bills list and edit it
            for main_bill in bills:
                if main_bill['name'] == bill['name'] and main_bill['due_date'] == bill['due_date']:
                    edit_bill_details(main_bill)
                    break
        else:
            error_msg("Invalid bill number.")
    except ValueError:
        error_msg("Please enter a valid number.")

def edit_bill_details(bill):
    """Edit details of a specific bill."""
    print(f"\n--- Editing '{bill['name']}' ---")
    
    new_name = colored_input(f"Name [{bill['name']}]: ", Colors.PROMPT).strip()
    if new_name:
        bill['name'] = new_name

    new_due_date = colored_input(f"Due Date [{bill['due_date']}]: ", Colors.PROMPT).strip()
    if new_due_date:
        try:
            datetime.strptime(new_due_date, DATE_FORMAT)
            is_valid, error_msg_text = validate_future_date(new_due_date)
            if is_valid:
                bill['due_date'] = new_due_date
            else:
                error_msg(error_msg_text + " Keeping the original date.")
        except ValueError:
            error_msg("Invalid date format. Keeping the original date.")
            
    # Website with validation
    colored_print(f"Current website: {bill.get('web_page', 'Not provided')}", Colors.INFO)
    new_web_page = colored_input(f"Web Page [{bill.get('web_page', '')}]: ", Colors.PROMPT).strip()
    if new_web_page:
        if new_web_page.lower() == 'clear':
            bill['web_page'] = ""
            success_msg("Website cleared.")
        else:
            validated_url = validate_url(new_web_page)
            if validated_url is not None:
                bill['web_page'] = validated_url
                if validated_url != new_web_page:
                    success_msg(f"Website corrected to: {validated_url}")
            else:
                error_msg("Invalid URL format. Keeping the original website.")

    new_login_info = colored_input(f"Login Info [{bill.get('login_info', '')}]: ", Colors.PROMPT).strip()
    if new_login_info:
        bill['login_info'] = new_login_info

    # Show decrypted password for editing
    current_password = bill.get('password', '')
    if current_password and CRYPTOGRAPHY_AVAILABLE and password_encryption.fernet:
        current_password = password_encryption.decrypt_password(current_password)
    
    new_password = colored_input(f"Password [{current_password}]: ", Colors.PROMPT).strip()
    if new_password:
        bill['password'] = new_password

    paid_status = colored_input(f"Paid (yes/no) [{'yes' if bill.get('paid', False) else 'no'}]: ", Colors.PROMPT).strip().lower()
    if paid_status in ['yes', 'y']:
        bill['paid'] = True
    elif paid_status in ['no', 'n']:
        bill['paid'] = False

    # Billing cycle
    current_cycle = bill.get('billing_cycle', 'monthly')
    print(f"\nCurrent billing cycle: {current_cycle.title()}")
    change_cycle = colored_input("Change billing cycle? (yes/no): ", Colors.PROMPT).strip().lower()
    if change_cycle in ['yes', 'y']:
        new_billing_cycle = get_billing_cycle()
        if new_billing_cycle is not None:
            bill['billing_cycle'] = new_billing_cycle
            success_msg(f"Billing cycle updated to {new_billing_cycle}")

    # Reminder period
    current_reminder = bill.get('reminder_days', 7)
    print(f"\nCurrent reminder period: {current_reminder} days before due date")
    change_reminder = colored_input("Change reminder period? (yes/no): ", Colors.PROMPT).strip().lower()
    if change_reminder in ['yes', 'y']:
        new_reminder_days = get_valid_reminder_days("Enter new reminder days before due date (1-365)", bill.get('reminder_days', 7))
        if new_reminder_days is not None:
            bill['reminder_days'] = new_reminder_days
            success_msg(f"Reminder period updated to {new_reminder_days} days")

    save_bills()
    success_msg(f"Bill '{bill['name']}' updated successfully.")
    input("Press Enter to continue...")

def pay_bill_from_due_list_paginated(current_due_bills, paginator):
    """Pay a specific bill from the paginated due bills list."""
    if not current_due_bills:
        return
    
    try:
        choice = int(colored_input(f"Enter bill number to pay (1-{len(current_due_bills)}): ", Colors.PROMPT))
        if 1 <= choice <= len(current_due_bills):
            bill, days_diff = current_due_bills[choice - 1]
            
            # Find and pay the bill
            for main_bill in bills:
                if main_bill['name'] == bill['name'] and main_bill['due_date'] == bill['due_date']:
                    main_bill['paid'] = True
                    save_bills()
                    success_msg(f"Bill '{bill['name']}' marked as paid!")
                    break
        else:
            error_msg("Invalid bill number.")
    except ValueError:
        error_msg("Please enter a valid number.")
    
    input("Press Enter to continue...")

def bulk_pay_bills_from_due_list(due_bills):
    """Pay multiple bills from the due bills list."""
    if not due_bills:
        warning_msg("No bills to pay.")
        return
    
    choice = colored_input(f"\nPay all {len(due_bills)} due bills? (yes/no): ", Colors.PROMPT).strip().lower()
    
    if choice in ['yes', 'y']:
        paid_count = 0
        for bill, _ in due_bills:
            # Find and pay the bill
            for main_bill in bills:
                if main_bill['name'] == bill['name'] and main_bill['due_date'] == bill['due_date']:
                    if not main_bill.get('paid', False):
                        main_bill['paid'] = True
                        paid_count += 1
                    break
        
        save_bills()
        success_msg(f"Paid {paid_count} bills successfully!")
    else:
        info_msg("Bulk payment cancelled.")
    
    input("Press Enter to continue...")

# 11. Sort functions (missing from your code)
def sort_bills():
    """Sort bills by different criteria."""
    if not bills:
        error_msg("No bills found to sort.")
        input("Press Enter to continue...")
        return
    
    print("\n--- Sort Bills ---")
    print("Sort options:")
    print("1. 📅 Sort by due date (earliest first)")
    print("2. 📅 Sort by due date (latest first)")
    print("3. 🔤 Sort by name (A-Z)")
    print("4. 🔤 Sort by name (Z-A)")
    print("5. ✅ Sort by payment status (unpaid first)")
    print("6. ✅ Sort by payment status (paid first)")
    print("7. 🔄 Reset to original order")
    print("8. 🚪 Back to main menu")
    
    choice = input("\nChoose sort option (1-8): ").strip()
    
    if choice == '1':
        sort_by_due_date_asc()
    elif choice == '2':
        sort_by_due_date_desc()
    elif choice == '3':
        sort_by_name_asc()
    elif choice == '4':
        sort_by_name_desc()
    elif choice == '5':
        sort_by_status_unpaid_first()
    elif choice == '6':
        sort_by_status_paid_first()
    elif choice == '7':
        reset_bill_order()
    elif choice == '8':
        return
    else:
        error_msg("Invalid option. Please choose 1-8.")
        input("Press Enter to continue...")
        sort_bills()

def sort_by_due_date_asc():
    """Sort bills by due date (earliest first)."""
    global bills
    try:
        bills.sort(key=lambda bill: datetime.strptime(bill['due_date'], DATE_FORMAT))
        success_msg("Bills sorted by due date (earliest first)")
        display_sorted_bills("Bills Sorted by Due Date (Earliest First)")
    except ValueError:
        error_msg("Invalid date format found")
        input("Press Enter to continue...")

def sort_by_due_date_desc():
    """Sort bills by due date (latest first)."""
    global bills
    try:
        bills.sort(key=lambda bill: datetime.strptime(bill['due_date'], DATE_FORMAT), reverse=True)
        success_msg("Bills sorted by due date (latest first)")
        display_sorted_bills("Bills Sorted by Due Date (Latest First)")
    except ValueError:
        error_msg("Invalid date format found")
        input("Press Enter to continue...")

def sort_by_name_asc():
    """Sort bills by name (A-Z)."""
    global bills
    bills.sort(key=lambda bill: bill['name'].lower())
    success_msg("Bills sorted by name (A-Z)")
    display_sorted_bills("Bills Sorted by Name (A-Z)")

def sort_by_name_desc():
    """Sort bills by name (Z-A)."""
    global bills
    bills.sort(key=lambda bill: bill['name'].lower(), reverse=True)
    success_msg("Bills sorted by name (Z-A)")
    display_sorted_bills("Bills Sorted by Name (Z-A)")

def sort_by_status_unpaid_first():
    """Sort bills by payment status (unpaid first)."""
    global bills
    bills.sort(key=lambda bill: bill.get('paid', False))
    success_msg("Bills sorted by status (unpaid first)")
    display_sorted_bills("Bills Sorted by Status (Unpaid First)")

def sort_by_status_paid_first():
    """Sort bills by payment status (paid first)."""
    global bills
    bills.sort(key=lambda bill: bill.get('paid', False), reverse=True)
    success_msg("Bills sorted by status (paid first)")
    display_sorted_bills("Bills Sorted by Status (Paid First)")

def reset_bill_order():
    """Reset bills to original order (reload from file)."""
    global bills
    load_bills()
    success_msg("Bills reset to original order")
    display_sorted_bills("Bills in Original Order")

def display_sorted_bills(title):
    """Display sorted bills with formatting."""
    print(f"\n--- {title} ---")
    
    if not bills:
        error_msg("No bills to display.")
        input("Press Enter to continue...")
        return
    
    for idx, bill in enumerate(bills, 1):
        status = "✓ Paid" if bill.get('paid', False) else "○ Unpaid"
        
        # Add due date info for better context
        try:
            due_date = datetime.strptime(bill['due_date'], DATE_FORMAT)
            today = datetime.now()
            days_diff = (due_date - today).days
            
            if days_diff < 0:
                date_info = f"(Overdue by {abs(days_diff)} days)"
            elif days_diff == 0:
                date_info = "(Due today!)"
            elif days_diff <= 7:
                date_info = f"(Due in {days_diff} days)"
            else:
                date_info = ""
        except ValueError:
            date_info = ""
        
        print(f"{idx:2}. {bill['name']} [{status}]")
        print(f"    Due: {bill['due_date']} {date_info}")
        if bill.get('web_page'):
            print(f"    Website: {bill['web_page']}")
        print()
    
    # Options after viewing sorted bills
    print("Options:")
    print("1. 💾 Save this sort order permanently")
    print("2. 🔄 Sort again with different criteria")
    print("3. 🚪 Back to main menu")
    
    choice = input("Choose option (1-3): ").strip()
    
    if choice == '1':
        save_bills()
        success_msg("Sort order saved!")
        input("Press Enter to continue...")
    elif choice == '2':
        sort_bills()
    elif choice == '3':
        return
    else:
        error_msg("Invalid option.")
        input("Press Enter to continue...")

def migrate_bills_to_billing_cycles():
    """Add billing cycle to existing bills that don't have it."""
    migrated_count = 0
    for bill in bills:
        if 'billing_cycle' not in bill:
            bill['billing_cycle'] = BillingCycle.MONTHLY  # Default to monthly
            migrated_count += 1
    
    if migrated_count > 0:
        save_bills()
        info_msg(f"Migrated {migrated_count} bills to include billing cycles (defaulted to monthly)")

def show_billing_cycle_summary():
    """Show a summary of bills by billing cycle."""
    if not bills:
        warning_msg("No bills found.")
        return
    
    title_msg("Bills by Billing Cycle")
    
    # Group bills by cycle
    cycle_groups = {}
    for bill in bills:
        cycle = bill.get('billing_cycle', BillingCycle.MONTHLY)
        if cycle not in cycle_groups:
            cycle_groups[cycle] = []
        cycle_groups[cycle].append(bill)
    
    # Display each group
    for cycle in BillingCycle.get_all_cycles():
        if cycle in cycle_groups:
            cycle_color = get_billing_cycle_color(cycle)
            bills_in_cycle = cycle_groups[cycle]
            
            print(f"\n{cycle_color}📅 {cycle.title()} ({len(bills_in_cycle)} bills){Colors.RESET}")
            print(f"   {BillingCycle.get_cycle_description(cycle)}")
            
            for bill in bills_in_cycle:
                status = "✓ Paid" if bill.get('paid', False) else "○ Unpaid"
                print(f"   • {bill['name']} - Due: {bill['due_date']} [{status}]")
    
    input("\nPress Enter to continue...")

def calculate_upcoming_bills(days=30):
    """Calculate all upcoming bill occurrences within the specified days."""
    upcoming = []
    today = datetime.now()
    end_date = today + timedelta(days=days)
    
    for bill in bills:
        if bill.get('billing_cycle') == BillingCycle.ONE_TIME and bill.get('paid', False):
            continue  # Skip completed one-time bills
        
        try:
            current_due = datetime.strptime(bill['due_date'], DATE_FORMAT)
            cycle = bill.get('billing_cycle', BillingCycle.MONTHLY)
            
            # Generate occurrences for this billing cycle
            occurrence_date = current_due
            occurrences = 0
            max_occurrences = 20  # Prevent infinite loops
            
            while occurrence_date <= end_date and occurrences < max_occurrences:
                if occurrence_date >= today:
                    days_until = (occurrence_date - today).days
                    upcoming.append({
                        'bill': bill,
                        'due_date': occurrence_date,
                        'days_until': days_until,
                        'cycle': cycle
                    })
                
                # Calculate next occurrence
                if cycle == BillingCycle.ONE_TIME:
                    break  # One-time bills don't repeat
                
                next_due_str = calculate_next_due_date(
                    occurrence_date.strftime(DATE_FORMAT), 
                    cycle
                )
                occurrence_date = datetime.strptime(next_due_str, DATE_FORMAT)
                occurrences += 1
                
        except ValueError:
            continue  # Skip bills with invalid dates
    
    return sorted(upcoming, key=lambda x: x['days_until'])

def show_upcoming_bills_calendar():
    """Show upcoming bills in a calendar-like view."""
    title_msg("Upcoming Bills Calendar (Next 30 Days)")
    
    upcoming = calculate_upcoming_bills(30)
    
    if not upcoming:
        info_msg("No upcoming bills in the next 30 days!")
        input("Press Enter to continue...")
        return
    
    # Group by weeks
    current_week = []
    current_week_start = None
    
    for occurrence in upcoming:
        due_date = occurrence['due_date']
        
        # Start new week if needed
        week_start = due_date - timedelta(days=due_date.weekday())
        if current_week_start != week_start:
            if current_week:
                display_week(current_week_start, current_week)
            current_week = []
            current_week_start = week_start
        
        current_week.append(occurrence)
    
    # Display last week
    if current_week:
        display_week(current_week_start, current_week)
    
    input("\nPress Enter to continue...")

def display_week(week_start, occurrences):
    """Display a week of bill occurrences."""
    week_end = week_start + timedelta(days=6)
    print(f"\n📅 Week of {week_start.strftime('%B %d')} - {week_end.strftime('%B %d, %Y')}")
    print("=" * 60)
    
    # Group occurrences by day
    days_dict = {}
    for occurrence in occurrences:
        day_key = occurrence['due_date'].strftime('%Y-%m-%d')
        if day_key not in days_dict:
            days_dict[day_key] = []
        days_dict[day_key].append(occurrence)
    
    # Display each day of the week
    for i in range(7):
        day = week_start + timedelta(days=i)
        day_key = day.strftime('%Y-%m-%d')
        day_name = day.strftime('%A')
        day_date = day.strftime('%m/%d')
        
        if day_key in days_dict:
            day_occurrences = days_dict[day_key]
            print(f"\n{Colors.INFO}{day_name} {day_date}:{Colors.RESET}")
            
            for occurrence in day_occurrences:
                bill = occurrence['bill']
                cycle = occurrence['cycle']
                cycle_color = get_billing_cycle_color(cycle)
                
                days_until = occurrence['days_until']
                if days_until == 0:
                    urgency = f"{Colors.DUE_SOON}DUE TODAY!{Colors.RESET}"
                elif days_until <= 3:
                    urgency = f"{Colors.WARNING}Due in {days_until} days{Colors.RESET}"
                else:
                    urgency = f"{Colors.INFO}Due in {days_until} days{Colors.RESET}"
                
                print(f"  • {bill['name']} ({cycle_color}{cycle}{Colors.RESET}) - {urgency}")


# Add to menu system
def billing_cycle_menu():
    """Show billing cycle management menu."""
    while True:
        clear_console()
        title_msg("Billing Cycle Management")
        
        print(f"{Colors.MENU}1.{Colors.RESET} 📊 Show bills by billing cycle")
        print(f"{Colors.MENU}2.{Colors.RESET} 📅 Show upcoming bills calendar")
        print(f"{Colors.MENU}3.{Colors.RESET} 🔄 Migrate existing bills to billing cycles")
        print(f"{Colors.MENU}4.{Colors.RESET} 🚪 Back to main menu")
        
        choice = colored_input("\nChoose option (1-4): ", Colors.PROMPT).strip()
        
        if choice == '1':
            clear_console()
            show_billing_cycle_summary()
        elif choice == '2':
            clear_console()
            show_upcoming_bills_calendar()
        elif choice == '3':
            clear_console()
            migrate_bills_to_billing_cycles()
            input("Press Enter to continue...")
        elif choice == '4':
            break
        else:
            error_msg("Invalid option. Please choose 1-4.")
            input("Press Enter to continue...")

# 5.1 Template operations
def load_templates():
    """Load bill templates from SQLite database."""
    global bill_templates
    try:
        # Initialize database if it doesn't exist
        initialize_database()
        
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM templates ORDER BY name')
        rows = cursor.fetchall()
        
        bill_templates = []
        for row in rows:
            template = dict(row)
            bill_templates.append(template)
        
        conn.close()
        success_msg(f"Loaded {len(bill_templates)} bill templates from database")
        
    except Exception as e:
        error_msg(f"Error loading templates from database: {e}")
        bill_templates = []

def save_templates():
    """Save bill templates to SQLite database."""
    try:
        # Initialize database if it doesn't exist
        initialize_database()
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Clear existing templates
        cursor.execute('DELETE FROM templates')
        
        # Insert all templates
        for template in bill_templates:
            cursor.execute('''
                INSERT INTO templates (
                    name, due_date, billing_cycle, reminder_days, web_page,
                    login_info, password, company_email, support_phone,
                    billing_phone, customer_service_hours, account_number,
                    reference_id, support_chat_url, mobile_app
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                template.get('name', ''),
                template.get('due_date', ''),
                template.get('billing_cycle', 'monthly'),
                template.get('reminder_days', 7),
                template.get('web_page', ''),
                template.get('login_info', ''),
                template.get('password', ''),
                template.get('company_email', ''),
                template.get('support_phone', ''),
                template.get('billing_phone', ''),
                template.get('customer_service_hours', ''),
                template.get('account_number', ''),
                template.get('reference_id', ''),
                template.get('support_chat_url', ''),
                template.get('mobile_app', '')
            ))
        
        conn.commit()
        conn.close()
        success_msg("Templates saved to database successfully")
        
    except Exception as e:
        error_msg(f"Template save error: {e}")

def create_template_from_bill(bill):
    """Create a template from an existing bill."""
    template = {
        "name": bill['name'],
        "web_page": bill.get('web_page', ''),
        "login_info": bill.get('login_info', ''),
        "password": bill.get('password', ''),
        "billing_cycle": bill.get('billing_cycle', BillingCycle.MONTHLY),
        "reminder_days": bill.get('reminder_days', 7),
        # Contact information
        "company_email": bill.get('company_email', ''),
        "support_phone": bill.get('support_phone', ''),
        "billing_phone": bill.get('billing_phone', ''),
        "customer_service_hours": bill.get('customer_service_hours', ''),
        "account_number": bill.get('account_number', ''),
        "reference_id": bill.get('reference_id', ''),
        "support_chat_url": bill.get('support_chat_url', ''),
        "mobile_app": bill.get('mobile_app', '')
    }
    return template

def save_bill_as_template():
    """Save an existing bill as a template."""
    if not bills:
        warning_msg("No bills found to create templates from.")
        return
    
    title_msg("Save Bill as Template")
    
    # Show bills for selection
    print(f"\n{Colors.INFO}📋 Available bills:{Colors.RESET}")
    for i, bill in enumerate(bills, 1):
        status = "✓ Paid" if bill.get('paid', False) else "○ Unpaid"
        print(f"{Colors.INFO}  {i}. {bill['name']} [{status}]{Colors.RESET}")
    
    try:
        choice = int(colored_input(f"\nSelect bill to save as template (1-{len(bills)}): ", Colors.PROMPT))
        if 1 <= choice <= len(bills):
            bill = bills[choice - 1]
            
            # Check if template already exists
            existing_template = next((t for t in bill_templates if t['name'] == bill['name']), None)
            if existing_template:
                overwrite = colored_input(f"Template '{bill['name']}' already exists. Overwrite? (y/n): ", Colors.WARNING).strip().lower()
                if overwrite not in ['y', 'yes']:
                    info_msg("Template creation cancelled.")
                    return
                bill_templates.remove(existing_template)
            
            # Create and save template
            template = create_template_from_bill(bill)
            bill_templates.append(template)
            save_templates()
            
            success_msg(f"Template '{bill['name']}' saved successfully!")
            
            # Show template details
            print(f"\n{Colors.INFO}📋 Template details:{Colors.RESET}")
            print(f"  Name: {Colors.TITLE}{template['name']}{Colors.RESET}")
            print(f"  Billing Cycle: {get_billing_cycle_color(template['billing_cycle'])}{template['billing_cycle'].title()}{Colors.RESET}")
            print(f"  Reminder Days: {Colors.INFO}{template['reminder_days']}{Colors.RESET}")
            if template.get('web_page'):
                print(f"  Website: {Colors.INFO}{template['web_page']}{Colors.RESET}")
            if template.get('login_info'):
                print(f"  Login: {Colors.INFO}{template['login_info']}{Colors.RESET}")
        else:
            error_msg("Invalid selection.")
    except ValueError:
        error_msg("Please enter a valid number.")
    
    colored_input("\nPress Enter to continue...", Colors.INFO)

def create_template_manually():
    """Create a new template manually."""
    title_msg("Create New Bill Template")
    info_msg("Type 'cancel' at any time to cancel.")
    
    # Get template name
    while True:
        name = get_required_input("Enter template name")
        if name is None:
            warning_msg("Template creation cancelled.")
            return
        
        # Check for duplicates
        if any(template['name'].lower() == name.lower() for template in bill_templates):
            error_msg(f"A template with the name '{name}' already exists. Please enter a different name.")
        else:
            break
    
    # Get billing cycle
    billing_cycle = get_billing_cycle()
    if billing_cycle is None:
        warning_msg("Template creation cancelled.")
        return
    
    # Get reminder period
    reminder_days = get_valid_reminder_days("Enter reminder days before due date (1-365)")
    if reminder_days is None:
        warning_msg("Template creation cancelled.")
        return
    
    # Get optional fields
    web_page = get_valid_url("Enter the web page for the bill (optional)")
    if web_page is None:
        warning_msg("Template creation cancelled.")
        return
    
    login_info = get_optional_input("Enter the login information for the bill")
    if login_info is None:
        warning_msg("Template creation cancelled.")
        return

    password = get_optional_input("Enter the password for the bill")
    if password is None:
        warning_msg("Template creation cancelled.")
        return

    # Get contact information
    print(f"\n{Colors.TITLE}📞 Contact Information (Optional){Colors.RESET}")
    print(f"{Colors.INFO}Add customer service contact details for this template:{Colors.RESET}")
    
    company_email = get_valid_email("Enter company customer service email: ")
    if company_email is None:
        warning_msg("Template creation cancelled.")
        return
    
    support_phone = get_optional_input("Enter customer support phone number")
    if support_phone is None:
        warning_msg("Template creation cancelled.")
        return
    
    billing_phone = get_optional_input("Enter billing department phone number")
    if billing_phone is None:
        warning_msg("Template creation cancelled.")
        return
    
    customer_service_hours = get_optional_input("Enter customer service hours (e.g., Mon-Fri 9AM-5PM)")
    if customer_service_hours is None:
        warning_msg("Template creation cancelled.")
        return
    
    account_number = get_optional_input("Enter account/customer number")
    if account_number is None:
        warning_msg("Template creation cancelled.")
        return
    
    reference_id = get_optional_input("Enter reference/policy number")
    if reference_id is None:
        warning_msg("Template creation cancelled.")
        return
    
    support_chat_url = get_valid_url("Enter live chat support URL (optional)")
    if support_chat_url is None:
        warning_msg("Template creation cancelled.")
        return
    
    mobile_app = get_optional_input("Enter mobile app information (e.g., 'Netflix App - iOS/Android')")
    if mobile_app is None:
        warning_msg("Template creation cancelled.")
        return

    # Create and save template
    template = {
        "name": name,
        "web_page": web_page,
        "login_info": login_info,
        "password": password,
        "billing_cycle": billing_cycle,
        "reminder_days": reminder_days,
        # Contact information
        "company_email": company_email,
        "support_phone": support_phone,
        "billing_phone": billing_phone,
        "customer_service_hours": customer_service_hours,
        "account_number": account_number,
        "reference_id": reference_id,
        "support_chat_url": support_chat_url,
        "mobile_app": mobile_app
    }
    
    bill_templates.append(template)
    save_templates()
    
    success_msg(f"Template '{name}' created successfully!")
    colored_input("Press Enter to continue...", Colors.INFO)

def view_templates():
    """View all available templates."""
    title_msg("Bill Templates")
    
    if not bill_templates:
        warning_msg("No templates found.")
        colored_input("Press Enter to continue...", Colors.INFO)
        return
    
    for idx, template in enumerate(bill_templates, 1):
        print(f"\n{Colors.INFO}{idx}.{Colors.RESET} {Colors.TITLE}{template['name']}{Colors.RESET}")
        
        # Show billing cycle
        cycle = template.get('billing_cycle', BillingCycle.MONTHLY)
        cycle_color = get_billing_cycle_color(cycle)
        print(f"    Cycle: {cycle_color}{cycle.title()}{Colors.RESET}")
        
        # Show reminder period
        reminder_days = template.get('reminder_days', 7)
        if reminder_days == 1:
            reminder_text = "1 day before"
        else:
            reminder_text = f"{reminder_days} days before"
        print(f"    Reminder: {Colors.WARNING}⏰ {reminder_text}{Colors.RESET}")
        
        if template.get('web_page'):
            print(f"    Website: {Colors.INFO}{template['web_page']}{Colors.RESET}")
        if template.get('login_info'):
            print(f"    Login: {Colors.INFO}{template['login_info']}{Colors.RESET}")
        
        # Show contact information if available
        contact_info = []
        if template.get('company_email'):
            contact_info.append(f"📧 {template['company_email']}")
        if template.get('support_phone'):
            contact_info.append(f"📞 {template['support_phone']}")
        if template.get('billing_phone'):
            contact_info.append(f"💰 {template['billing_phone']}")
        if template.get('account_number'):
            contact_info.append(f"🆔 {template['account_number']}")
        
        if contact_info:
            print(f"    {Colors.INFO}📞 Contact: {', '.join(contact_info[:2])}{'...' if len(contact_info) > 2 else ''}{Colors.RESET}")
    
    # Template management options
    print(f"\n{Colors.MENU}Template Options:{Colors.RESET}")
    print("1. ✏️  Edit a template")
    print("2. 🗑️  Delete a template")
    print("3. 📝 Use template to add bill")
    print("4. 🔙 Back to templates menu")
    
    choice = colored_input("\nChoose option (1-4): ", Colors.PROMPT).strip()
    
    if choice == '1':
        edit_template()
    elif choice == '2':
        delete_template()
    elif choice == '3':
        use_template_to_add_bill()
    elif choice == '4':
        return
    else:
        error_msg("Invalid option.")
        colored_input("Press Enter to continue...", Colors.WARNING)

def edit_template():
    """Edit an existing template."""
    if not bill_templates:
        warning_msg("No templates to edit.")
        return
    
    try:
        choice = int(colored_input(f"Enter template number to edit (1-{len(bill_templates)}): ", Colors.PROMPT))
        if 1 <= choice <= len(bill_templates):
            template = bill_templates[choice - 1]
            print(f"\n--- Editing Template: {template['name']} ---")
            
            # Edit template fields
            new_name = colored_input(f"Name [{template['name']}]: ", Colors.PROMPT).strip()
            if new_name:
                # Check for name conflicts
                if any(t['name'].lower() == new_name.lower() and t != template for t in bill_templates):
                    error_msg(f"A template with the name '{new_name}' already exists.")
                else:
                    template['name'] = new_name
            
            # Billing cycle
            current_cycle = template.get('billing_cycle', BillingCycle.MONTHLY)
            print(f"\nCurrent billing cycle: {current_cycle.title()}")
            change_cycle = colored_input("Change billing cycle? (yes/no): ", Colors.PROMPT).strip().lower()
            if change_cycle in ['yes', 'y']:
                new_billing_cycle = get_billing_cycle()
                if new_billing_cycle is not None:
                    template['billing_cycle'] = new_billing_cycle
            
            # Reminder period
            current_reminder = template.get('reminder_days', 7)
            print(f"\nCurrent reminder period: {current_reminder} days before due date")
            change_reminder = colored_input("Change reminder period? (yes/no): ", Colors.PROMPT).strip().lower()
            if change_reminder in ['yes', 'y']:
                new_reminder_days = get_valid_reminder_days("Enter new reminder days before due date (1-365)", current_reminder)
                if new_reminder_days is not None:
                    template['reminder_days'] = new_reminder_days
            
            # Website
            new_web_page = colored_input(f"Website [{template.get('web_page', '')}]: ", Colors.PROMPT).strip()
            if new_web_page:
                if new_web_page.lower() == 'clear':
                    template['web_page'] = ""
                else:
                    validated_url = validate_url(new_web_page)
                    if validated_url is not None:
                        template['web_page'] = validated_url
                    else:
                        error_msg("Invalid URL format. Keeping the original website.")
            
            # Login info
            new_login_info = colored_input(f"Login Info [{template.get('login_info', '')}]: ", Colors.PROMPT).strip()
            if new_login_info:
                template['login_info'] = new_login_info
            
            # Password
            # Show decrypted password for editing
            current_password = template.get('password', '')
            if current_password and CRYPTOGRAPHY_AVAILABLE and password_encryption.fernet:
                current_password = password_encryption.decrypt_password(current_password)
            
            new_password = colored_input(f"Password [{current_password}]: ", Colors.PROMPT).strip()
            if new_password:
                template['password'] = new_password
            
            # Contact Information
            print(f"\n{Colors.TITLE}📞 Contact Information{Colors.RESET}")
            
            new_company_email = colored_input(f"Company Email [{template.get('company_email', '')}]: ", Colors.PROMPT).strip()
            if new_company_email:
                if new_company_email.lower() == 'clear':
                    template['company_email'] = ""
                    success_msg("Company email cleared.")
                else:
                    validated_email = validate_email(new_company_email)
                    if validated_email is not None:
                        template['company_email'] = validated_email
                    else:
                        error_msg("Invalid email format. Keeping the original email.")

            new_support_phone = colored_input(f"Support Phone [{template.get('support_phone', '')}]: ", Colors.PROMPT).strip()
            if new_support_phone:
                template['support_phone'] = new_support_phone

            new_billing_phone = colored_input(f"Billing Phone [{template.get('billing_phone', '')}]: ", Colors.PROMPT).strip()
            if new_billing_phone:
                template['billing_phone'] = new_billing_phone

            new_service_hours = colored_input(f"Service Hours [{template.get('customer_service_hours', '')}]: ", Colors.PROMPT).strip()
            if new_service_hours:
                template['customer_service_hours'] = new_service_hours

            new_account_number = colored_input(f"Account Number [{template.get('account_number', '')}]: ", Colors.PROMPT).strip()
            if new_account_number:
                template['account_number'] = new_account_number

            new_reference_id = colored_input(f"Reference ID [{template.get('reference_id', '')}]: ", Colors.PROMPT).strip()
            if new_reference_id:
                template['reference_id'] = new_reference_id

            new_support_chat_url = colored_input(f"Support Chat URL [{template.get('support_chat_url', '')}]: ", Colors.PROMPT).strip()
            if new_support_chat_url:
                if new_support_chat_url.lower() == 'clear':
                    template['support_chat_url'] = ""
                    success_msg("Support chat URL cleared.")
                else:
                    validated_url = validate_url(new_support_chat_url)
                    if validated_url is not None:
                        template['support_chat_url'] = validated_url
                    else:
                        error_msg("Invalid URL format. Keeping the original URL.")

            new_mobile_app = colored_input(f"Mobile App [{template.get('mobile_app', '')}]: ", Colors.PROMPT).strip()
            if new_mobile_app:
                template['mobile_app'] = new_mobile_app
            
            save_templates()
            success_msg(f"Template '{template['name']}' updated successfully.")
        else:
            error_msg("Invalid selection.")
    except ValueError:
        error_msg("Please enter a valid number.")
    
    colored_input("Press Enter to continue...", Colors.INFO)

def delete_template():
    """Delete a template."""
    if not bill_templates:
        warning_msg("No templates to delete.")
        return
    
    try:
        choice = int(colored_input(f"Enter template number to delete (1-{len(bill_templates)}): ", Colors.PROMPT))
        if 1 <= choice <= len(bill_templates):
            template = bill_templates[choice - 1]
            confirm = colored_input(f"Are you sure you want to delete template '{template['name']}'? (yes/no): ", Colors.WARNING).strip().lower()
            if confirm in ['yes', 'y']:
                bill_templates.remove(template)
                save_templates()
                success_msg(f"Template '{template['name']}' deleted successfully.")
            else:
                info_msg("Deletion cancelled.")
        else:
            error_msg("Invalid selection.")
    except ValueError:
        error_msg("Please enter a valid number.")
    
    colored_input("Press Enter to continue...", Colors.INFO)

def use_template_to_add_bill():
    """Use a template to quickly add a new bill."""
    if not bill_templates:
        warning_msg("No templates available.")
        colored_input("Press Enter to continue...", Colors.INFO)
        return
    
    try:
        choice = int(colored_input(f"Enter template number to use (1-{len(bill_templates)}): ", Colors.PROMPT))
        if 1 <= choice <= len(bill_templates):
            template = bill_templates[choice - 1]
            
            title_msg(f"Add Bill from Template: {template['name']}")
            
            # Get due date (required for new bill)
            due_date = get_valid_date_with_range_check("Enter the due date for this bill (YYYY-MM-DD)")
            if due_date is None:
                warning_msg("Bill creation cancelled.")
                return
            
            # Create bill from template
            bill = {
                "name": template['name'],
                "due_date": due_date,
                "web_page": template.get('web_page', ''),
                "login_info": template.get('login_info', ''),
                "password": template.get('password', ''),
                "paid": False,
                "billing_cycle": template.get('billing_cycle', BillingCycle.MONTHLY),
                "reminder_days": template.get('reminder_days', 7),
                # Contact information from template
                "company_email": template.get('company_email', ''),
                "support_phone": template.get('support_phone', ''),
                "billing_phone": template.get('billing_phone', ''),
                "customer_service_hours": template.get('customer_service_hours', ''),
                "account_number": template.get('account_number', ''),
                "reference_id": template.get('reference_id', ''),
                "support_chat_url": template.get('support_chat_url', ''),
                "mobile_app": template.get('mobile_app', '')
            }
            
            bills.append(bill)
            save_bills()
            
            success_msg(f"Bill '{template['name']}' added successfully from template!")
            info_msg(f"Due date: {due_date}")
            info_msg(f"Billing cycle: {template.get('billing_cycle', BillingCycle.MONTHLY)}")
            
            # Show contact info if available
            contact_info = []
            if template.get('company_email'):
                contact_info.append(f"📧 {template['company_email']}")
            if template.get('support_phone'):
                contact_info.append(f"📞 {template['support_phone']}")
            if template.get('account_number'):
                contact_info.append(f"🆔 {template['account_number']}")
            
            if contact_info:
                info_msg(f"Contact info included: {', '.join(contact_info[:2])}{'...' if len(contact_info) > 2 else ''}")
            
        else:
            error_msg("Invalid selection.")
    except ValueError:
        error_msg("Please enter a valid number.")
    
    colored_input("Press Enter to continue...", Colors.INFO)

def templates_menu():
    """Display templates management menu."""
    while True:
        clear_console()
        title_msg("Bill Templates Management")
        
        print(f"{Colors.MENU}1.{Colors.RESET} 📋 View all templates")
        print(f"{Colors.MENU}2.{Colors.RESET} 📝 Create new template")
        print(f"{Colors.MENU}3.{Colors.RESET} 💾 Save bill as template")
        print(f"{Colors.MENU}4.{Colors.RESET} 🚪 Back to main menu")
        
        choice = colored_input("\nChoose option (1-4): ", Colors.PROMPT).strip()
        
        if choice == '1':
            clear_console()
            view_templates()
        elif choice == '2':
            clear_console()
            create_template_manually()
        elif choice == '3':
            clear_console()
            save_bill_as_template()
        elif choice == '4':
            break
        else:
            error_msg("Invalid option. Please choose 1-4.")
            colored_input("Press Enter to continue...", Colors.WARNING)

# 5.2 CSV Import/Export operations
def import_bills_from_csv():
    """Import bills from a CSV file."""
    title_msg("Import Bills from CSV")
    info_msg("This will import bills from a CSV file. Make sure your CSV has the correct format.")
    
    # Show CSV format requirements
    print(f"\n{Colors.TITLE}📋 Required CSV Format:{Colors.RESET}")
    print("Your CSV file should have these columns (headers are case-insensitive):")
    print(f"{Colors.INFO}• name{Colors.RESET} - Bill name (required)")
    print(f"{Colors.INFO}• due_date{Colors.RESET} - Due date in YYYY-MM-DD format (required)")
    print(f"{Colors.INFO}• billing_cycle{Colors.RESET} - weekly, bi-weekly, monthly, quarterly, semi-annually, annually, one-time")
    print(f"{Colors.INFO}• reminder_days{Colors.RESET} - Days before due date for reminders (default: 7)")
    print(f"{Colors.INFO}• web_page{Colors.RESET} - Website URL (optional)")
    print(f"{Colors.INFO}• login_info{Colors.RESET} - Login information (optional)")
    print(f"{Colors.INFO}• password{Colors.RESET} - Password (optional)")
    print(f"{Colors.INFO}• company_email{Colors.RESET} - Customer service email (optional)")
    print(f"{Colors.INFO}• support_phone{Colors.RESET} - Support phone number (optional)")
    print(f"{Colors.INFO}• billing_phone{Colors.RESET} - Billing phone number (optional)")
    print(f"{Colors.INFO}• customer_service_hours{Colors.RESET} - Service hours (optional)")
    print(f"{Colors.INFO}• account_number{Colors.RESET} - Account number (optional)")
    print(f"{Colors.INFO}• reference_id{Colors.RESET} - Reference ID (optional)")
    print(f"{Colors.INFO}• support_chat_url{Colors.RESET} - Live chat URL (optional)")
    print(f"{Colors.INFO}• mobile_app{Colors.RESET} - Mobile app info (optional)")
    
    # Get CSV file path
    csv_file = colored_input(f"\n{Colors.PROMPT}Enter the path to your CSV file: {Colors.RESET}").strip()
    
    if not csv_file:
        warning_msg("Import cancelled.")
        return
    
    # Check if file exists
    if not os.path.exists(csv_file):
        error_msg(f"File '{csv_file}' not found.")
        colored_input("Press Enter to continue...", Colors.INFO)
        return
    
    # Validate file extension
    if not csv_file.lower().endswith('.csv'):
        error_msg("File must have a .csv extension.")
        colored_input("Press Enter to continue...", Colors.INFO)
        return
    
    try:
        # Read and parse CSV file
        imported_bills = []
        skipped_bills = []
        errors = []
        
        with open(csv_file, 'r', newline='', encoding='utf-8') as file:
            reader = csv.DictReader(file)
            
            # Validate headers
            required_headers = ['name', 'due_date']
            optional_headers = ['billing_cycle', 'reminder_days', 'web_page', 'login_info', 'password',
                              'company_email', 'support_phone', 'billing_phone', 'customer_service_hours',
                              'account_number', 'reference_id', 'support_chat_url', 'mobile_app']
            
            fieldnames = [field.lower() for field in reader.fieldnames] if reader.fieldnames else []
            
            missing_required = [h for h in required_headers if h not in fieldnames]
            if missing_required:
                error_msg(f"Missing required columns: {', '.join(missing_required)}")
                colored_input("Press Enter to continue...", Colors.INFO)
                return
            
            # Process each row
            row_count = 0
            for row in reader:
                row_count += 1
                try:
                    # Validate required fields
                    name = row.get('name', '').strip()
                    due_date = row.get('due_date', '').strip()
                    
                    if not name:
                        errors.append(f"Row {row_count}: Missing bill name")
                        continue
                    
                    if not due_date:
                        errors.append(f"Row {row_count}: Missing due date")
                        continue
                    
                    # Validate date format
                    try:
                        datetime.strptime(due_date, DATE_FORMAT)
                    except ValueError:
                        errors.append(f"Row {row_count}: Invalid date format '{due_date}' (use YYYY-MM-DD)")
                        continue
                    
                    # Check for duplicate names
                    if any(bill['name'].lower() == name.lower() for bill in bills):
                        skipped_bills.append(f"Row {row_count}: '{name}' (duplicate name)")
                        continue
                    
                    # Create bill object
                    bill_data = {
                        'name': name,
                        'due_date': due_date,
                        'paid': False,
                        'billing_cycle': row.get('billing_cycle', 'monthly').strip().lower(),
                        'reminder_days': int(row.get('reminder_days', 7)),
                        'web_page': row.get('web_page', '').strip(),
                        'login_info': row.get('login_info', '').strip(),
                        'password': row.get('password', '').strip(),
                        'company_email': row.get('company_email', '').strip(),
                        'support_phone': row.get('support_phone', '').strip(),
                        'billing_phone': row.get('billing_phone', '').strip(),
                        'customer_service_hours': row.get('customer_service_hours', '').strip(),
                        'account_number': row.get('account_number', '').strip(),
                        'reference_id': row.get('reference_id', '').strip(),
                        'support_chat_url': row.get('support_chat_url', '').strip(),
                        'mobile_app': row.get('mobile_app', '').strip()
                    }
                    
                    # Use comprehensive validation
                    is_valid, error_msg_text, cleaned_bill = DataValidator.validate_bill_data(bill_data)
                    if not is_valid:
                        errors.append(f"Row {row_count}: Validation failed - {error_msg_text}")
                        continue
                    
                    imported_bills.append(cleaned_bill)
                    
                except Exception as e:
                    errors.append(f"Row {row_count}: {str(e)}")
        
        # Show import results
        print(f"\n{Colors.TITLE}📊 Import Results:{Colors.RESET}")
        success_msg(f"Successfully imported {len(imported_bills)} bills")
        
        if skipped_bills:
            warning_msg(f"Skipped {len(skipped_bills)} bills (duplicates)")
            for skipped in skipped_bills[:5]:  # Show first 5
                print(f"  • {skipped}")
            if len(skipped_bills) > 5:
                print(f"  ... and {len(skipped_bills) - 5} more")
        
        if errors:
            error_msg(f"Found {len(errors)} errors")
            for error in errors[:5]:  # Show first 5
                print(f"  • {error}")
            if len(errors) > 5:
                print(f"  ... and {len(errors) - 5} more")
        
        # Ask user to confirm import
        if imported_bills:
            confirm = colored_input(f"\n{Colors.WARNING}Import {len(imported_bills)} bills? (yes/no): {Colors.RESET}").strip().lower()
            if confirm in ['yes', 'y']:
                # Add bills to the main list
                bills.extend(imported_bills)
                save_bills()
                success_msg(f"Successfully imported {len(imported_bills)} bills!")
                
                # Show sample of imported bills
                print(f"\n{Colors.INFO}📋 Sample of imported bills:{Colors.RESET}")
                for i, bill in enumerate(imported_bills[:3], 1):
                    print(f"  {i}. {bill['name']} - Due: {bill['due_date']} ({bill['billing_cycle']})")
                if len(imported_bills) > 3:
                    print(f"  ... and {len(imported_bills) - 3} more bills")
            else:
                info_msg("Import cancelled.")
        else:
            warning_msg("No bills to import.")
        
    except Exception as e:
        error_msg(f"Error reading CSV file: {str(e)}")
    
    colored_input("\nPress Enter to continue...", Colors.INFO)

def export_bills_to_csv():
    """Export bills to a CSV file."""
    title_msg("Export Bills to CSV")
    
    if not bills:
        warning_msg("No bills to export.")
        colored_input("Press Enter to continue...", Colors.INFO)
        return
    
    # Get export file path
    default_filename = f"bills_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    csv_file = colored_input(f"{Colors.PROMPT}Enter export filename [{default_filename}]: {Colors.RESET}").strip()
    
    if not csv_file:
        csv_file = default_filename
    
    # Ensure .csv extension
    if not csv_file.lower().endswith('.csv'):
        csv_file += '.csv'
    
    try:
        with open(csv_file, 'w', newline='', encoding='utf-8') as file:
            # Define fieldnames for CSV
            fieldnames = [
                'name', 'due_date', 'paid', 'billing_cycle', 'reminder_days',
                'web_page', 'login_info', 'password', 'company_email',
                'support_phone', 'billing_phone', 'customer_service_hours',
                'account_number', 'reference_id', 'support_chat_url', 'mobile_app'
            ]
            
            writer = csv.DictWriter(file, fieldnames=fieldnames)
            writer.writeheader()
            
            # Write each bill
            for bill in bills:
                row = {
                    'name': bill.get('name', ''),
                    'due_date': bill.get('due_date', ''),
                    'paid': 'yes' if bill.get('paid', False) else 'no',
                    'billing_cycle': bill.get('billing_cycle', 'monthly'),
                    'reminder_days': bill.get('reminder_days', 7),
                    'web_page': bill.get('web_page', ''),
                    'login_info': bill.get('login_info', ''),
                    'password': bill.get('password', ''),
                    'company_email': bill.get('company_email', ''),
                    'support_phone': bill.get('support_phone', ''),
                    'billing_phone': bill.get('billing_phone', ''),
                    'customer_service_hours': bill.get('customer_service_hours', ''),
                    'account_number': bill.get('account_number', ''),
                    'reference_id': bill.get('reference_id', ''),
                    'support_chat_url': bill.get('support_chat_url', ''),
                    'mobile_app': bill.get('mobile_app', '')
                }
                writer.writerow(row)
        
        success_msg(f"Successfully exported {len(bills)} bills to '{csv_file}'")
        info_msg(f"File location: {os.path.abspath(csv_file)}")
        
    except Exception as e:
        error_msg(f"Error exporting to CSV: {str(e)}")
    
    colored_input("Press Enter to continue...", Colors.INFO)

def create_sample_csv():
    """Create a sample CSV file with the correct format."""
    sample_filename = "sample_bills_import.csv"
    
    try:
        with open(sample_filename, 'w', newline='', encoding='utf-8') as file:
            fieldnames = [
                'name', 'due_date', 'billing_cycle', 'reminder_days',
                'web_page', 'login_info', 'password', 'company_email',
                'support_phone', 'billing_phone', 'customer_service_hours',
                'account_number', 'reference_id', 'support_chat_url', 'mobile_app'
            ]
            
            writer = csv.DictWriter(file, fieldnames=fieldnames)
            writer.writeheader()
            
            # Add sample data
            sample_bills = [
                {
                    'name': 'Netflix Subscription',
                    'due_date': '2024-02-15',
                    'billing_cycle': 'monthly',
                    'reminder_days': '7',
                    'web_page': 'https://netflix.com',
                    'login_info': 'user@example.com',
                    'password': 'your_password',
                    'company_email': 'support@netflix.com',
                    'support_phone': '1-800-123-4567',
                    'billing_phone': '1-800-123-4568',
                    'customer_service_hours': '24/7',
                    'account_number': 'NF123456789',
                    'reference_id': '',
                    'support_chat_url': 'https://netflix.com/help',
                    'mobile_app': 'Netflix App - iOS/Android'
                },
                {
                    'name': 'Electric Bill',
                    'due_date': '2024-02-20',
                    'billing_cycle': 'monthly',
                    'reminder_days': '10',
                    'web_page': 'https://electriccompany.com',
                    'login_info': 'account123',
                    'password': 'your_password',
                    'company_email': 'billing@electriccompany.com',
                    'support_phone': '1-800-555-0123',
                    'billing_phone': '1-800-555-0124',
                    'customer_service_hours': 'Mon-Fri 8AM-6PM',
                    'account_number': 'ELEC789012',
                    'reference_id': 'INV-2024-001',
                    'support_chat_url': '',
                    'mobile_app': 'Electric Company App'
                }
            ]
            
            for bill in sample_bills:
                writer.writerow(bill)
        
        success_msg(f"Sample CSV file created: '{sample_filename}'")
        info_msg(f"File location: {os.path.abspath(sample_filename)}")
        print(f"\n{Colors.INFO}📋 Sample file includes:{Colors.RESET}")
        print("  • Netflix Subscription (monthly)")
        print("  • Electric Bill (monthly)")
        print("  • All contact information fields")
        print("  • Proper date format (YYYY-MM-DD)")
        
    except Exception as e:
        error_msg(f"Error creating sample file: {str(e)}")
    
    colored_input("Press Enter to continue...", Colors.INFO)

def csv_import_export_menu():
    """Display CSV import/export menu."""
    while True:
        clear_console()
        title_msg("CSV Import/Export")
        
        print(f"{Colors.MENU}1.{Colors.RESET} 📥 Import bills from CSV")
        print(f"{Colors.MENU}2.{Colors.RESET} 📤 Export bills to CSV")
        print(f"{Colors.MENU}3.{Colors.RESET} 📋 Create sample CSV file")
        print(f"{Colors.MENU}4.{Colors.RESET} 📥 Import bills from Excel (.xlsx)")
        print(f"{Colors.MENU}5.{Colors.RESET} 📤 Export bills to Excel (.xlsx)")
        print(f"{Colors.MENU}6.{Colors.RESET} 📋 Create sample Excel file (.xlsx)")
        print(f"{Colors.MENU}7.{Colors.RESET} 🚪 Back to main menu")
        
        choice = colored_input("\nChoose option (1-7): ", Colors.PROMPT).strip()
        
        if choice == '1':
            clear_console()
            import_bills_from_csv()
        elif choice == '2':
            clear_console()
            export_bills_to_csv()
        elif choice == '3':
            clear_console()
            create_sample_csv()
        elif choice == '4':
            clear_console()
            import_bills_from_excel()
        elif choice == '5':
            clear_console()
            export_bills_to_excel()
        elif choice == '6':
            clear_console()
            create_sample_excel()
        elif choice == '7':
            break
        else:
            error_msg("Invalid option. Please choose 1-7.")
            colored_input("Press Enter to continue...", Colors.WARNING)

def import_bills_from_excel():
    """Import bills from an Excel (.xlsx) file using openpyxl and full validation."""
    from openpyxl import load_workbook
    title_msg("Import Bills from Excel (.xlsx)")
    info_msg("This will import bills from an Excel file. Make sure your file has the correct format.")
    
    # Show Excel format requirements (same as CSV)
    print(f"\n{Colors.TITLE}📋 Required Excel Format:{Colors.RESET}")
    print("Your Excel file should have these columns (headers are case-insensitive):")
    print(f"{Colors.INFO}• name{Colors.RESET} - Bill name (required)")
    print(f"{Colors.INFO}• due_date{Colors.RESET} - Due date in YYYY-MM-DD format (required)")
    print(f"{Colors.INFO}• billing_cycle{Colors.RESET} - weekly, bi-weekly, monthly, quarterly, semi-annually, annually, one-time")
    print(f"{Colors.INFO}• reminder_days{Colors.RESET} - Days before due date for reminders (default: 7)")
    print(f"{Colors.INFO}• web_page{Colors.RESET} - Website URL (optional)")
    print(f"{Colors.INFO}• login_info{Colors.RESET} - Login information (optional)")
    print(f"{Colors.INFO}• password{Colors.RESET} - Password (optional)")
    print(f"{Colors.INFO}• company_email{Colors.RESET} - Customer service email (optional)")
    print(f"{Colors.INFO}• support_phone{Colors.RESET} - Support phone number (optional)")
    print(f"{Colors.INFO}• billing_phone{Colors.RESET} - Billing phone number (optional)")
    print(f"{Colors.INFO}• customer_service_hours{Colors.RESET} - Service hours (optional)")
    print(f"{Colors.INFO}• account_number{Colors.RESET} - Account number (optional)")
    print(f"{Colors.INFO}• reference_id{Colors.RESET} - Reference ID (optional)")
    print(f"{Colors.INFO}• support_chat_url{Colors.RESET} - Live chat URL (optional)")
    print(f"{Colors.INFO}• mobile_app{Colors.RESET} - Mobile app info (optional)")
    
    # Get Excel file path
    xlsx_file = colored_input(f"\n{Colors.PROMPT}Enter the path to your Excel file: {Colors.RESET}").strip()
    if not xlsx_file:
        warning_msg("Import cancelled.")
        return
    if not os.path.exists(xlsx_file):
        error_msg(f"File '{xlsx_file}' not found.")
        colored_input("Press Enter to continue...", Colors.INFO)
        return
    if not xlsx_file.lower().endswith('.xlsx'):
        error_msg("File must have a .xlsx extension.")
        colored_input("Press Enter to continue...", Colors.INFO)
        return
    try:
        wb = load_workbook(xlsx_file)
        ws = wb.active
        # Read headers
        headers = [str(cell.value).strip().lower() if cell.value else '' for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        required_headers = ['name', 'due_date']
        missing_required = [h for h in required_headers if h not in headers]
        if missing_required:
            error_msg(f"Missing required columns: {', '.join(missing_required)}")
            colored_input("Press Enter to continue...", Colors.INFO)
            return
        # Read and validate rows
        imported_bills = []
        skipped_bills = []
        errors = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            row_data = {headers[i]: (cell.value if cell.value is not None else '') for i, cell in enumerate(row) if i < len(headers)}
            name = str(row_data.get('name', '')).strip()
            due_date = str(row_data.get('due_date', '')).strip()
            if not name:
                errors.append(f"Row {row_idx}: Missing bill name")
                continue
            if not due_date:
                errors.append(f"Row {row_idx}: Missing due date")
                continue
            # Validate date format
            try:
                datetime.strptime(due_date, DATE_FORMAT)
            except ValueError:
                errors.append(f"Row {row_idx}: Invalid date format '{due_date}' (use YYYY-MM-DD)")
                continue
            # Check for duplicate names
            if any(bill['name'].lower() == name.lower() for bill in bills):
                skipped_bills.append(f"Row {row_idx}: '{name}' (duplicate name)")
                continue
            # Build bill dict
            bill_data = {
                'name': name,
                'due_date': due_date,
                'paid': False,
                'billing_cycle': str(row_data.get('billing_cycle', 'monthly')).strip().lower(),
                'reminder_days': int(row_data.get('reminder_days', 7)) if str(row_data.get('reminder_days', '')).strip().isdigit() else 7,
                'web_page': str(row_data.get('web_page', '')).strip(),
                'login_info': str(row_data.get('login_info', '')).strip(),
                'password': str(row_data.get('password', '')).strip(),
                'company_email': str(row_data.get('company_email', '')).strip(),
                'support_phone': str(row_data.get('support_phone', '')).strip(),
                'billing_phone': str(row_data.get('billing_phone', '')).strip(),
                'customer_service_hours': str(row_data.get('customer_service_hours', '')).strip(),
                'account_number': str(row_data.get('account_number', '')).strip(),
                'reference_id': str(row_data.get('reference_id', '')).strip(),
                'support_chat_url': str(row_data.get('support_chat_url', '')).strip(),
                'mobile_app': str(row_data.get('mobile_app', '')).strip()
            }
            # Use comprehensive validation
            is_valid, error_msg_text, cleaned_bill = DataValidator.validate_bill_data(bill_data)
            if not is_valid:
                errors.append(f"Row {row_idx}: Validation failed - {error_msg_text}")
                continue
            imported_bills.append(cleaned_bill)
        # Show import results
        print(f"\n{Colors.TITLE}📊 Import Results:{Colors.RESET}")
        success_msg(f"Successfully imported {len(imported_bills)} bills")
        if skipped_bills:
            warning_msg(f"Skipped {len(skipped_bills)} bills (duplicates)")
            for skipped in skipped_bills[:5]:
                print(f"  • {skipped}")
            if len(skipped_bills) > 5:
                print(f"  ... and {len(skipped_bills) - 5} more")
        if errors:
            error_msg(f"Found {len(errors)} errors")
            for error in errors[:5]:
                print(f"  • {error}")
            if len(errors) > 5:
                print(f"  ... and {len(errors) - 5} more")
        # Ask user to confirm import
        if imported_bills:
            confirm = colored_input(f"\n{Colors.WARNING}Import {len(imported_bills)} bills? (yes/no): {Colors.RESET}").strip().lower()
            if confirm in ['yes', 'y']:
                bills.extend(imported_bills)
                save_bills()
                success_msg(f"Successfully imported {len(imported_bills)} bills!")
                print(f"\n{Colors.INFO}📋 Sample of imported bills:{Colors.RESET}")
                for i, bill in enumerate(imported_bills[:3], 1):
                    print(f"  {i}. {bill['name']} - Due: {bill['due_date']} ({bill['billing_cycle']})")
                if len(imported_bills) > 3:
                    print(f"  ... and {len(imported_bills) - 3} more bills")
            else:
                info_msg("Import cancelled.")
        else:
            warning_msg("No bills to import.")
    except Exception as e:
        error_msg(f"Error reading Excel file: {str(e)}")
    colored_input("\nPress Enter to continue...", Colors.INFO)

def export_bills_to_excel():
    """Export bills to an Excel (.xlsx) file using openpyxl."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    title_msg("Export Bills to Excel (.xlsx)")
    
    if not bills:
        warning_msg("No bills to export.")
        colored_input("Press Enter to continue...", Colors.INFO)
        return
    
    # Get export file path
    default_filename = f"bills_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    xlsx_file = colored_input(f"{Colors.PROMPT}Enter export filename [{default_filename}]: {Colors.RESET}").strip()
    if not xlsx_file:
        xlsx_file = default_filename
    if not xlsx_file.lower().endswith('.xlsx'):
        xlsx_file += '.xlsx'
    
    # Define columns/fields (same as CSV)
    fieldnames = [
        'name', 'due_date', 'paid', 'billing_cycle', 'reminder_days',
        'web_page', 'login_info', 'password', 'company_email',
        'support_phone', 'billing_phone', 'customer_service_hours',
        'account_number', 'reference_id', 'support_chat_url', 'mobile_app'
    ]
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Bills"
        # Write header
        ws.append(fieldnames)
        # Write each bill
        for bill in bills:
            row = [
                bill.get('name', ''),
                bill.get('due_date', ''),
                'yes' if bill.get('paid', False) else 'no',
                bill.get('billing_cycle', 'monthly'),
                bill.get('reminder_days', 7),
                bill.get('web_page', ''),
                bill.get('login_info', ''),
                bill.get('password', ''),
                bill.get('company_email', ''),
                bill.get('support_phone', ''),
                bill.get('billing_phone', ''),
                bill.get('customer_service_hours', ''),
                bill.get('account_number', ''),
                bill.get('reference_id', ''),
                bill.get('support_chat_url', ''),
                bill.get('mobile_app', '')
            ]
            ws.append(row)
        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max(12, min(max_length + 2, 40))
        wb.save(xlsx_file)
        success_msg(f"Successfully exported {len(bills)} bills to '{xlsx_file}'")
        info_msg(f"File location: {os.path.abspath(xlsx_file)}")
    except Exception as e:
        error_msg(f"Error exporting to Excel: {str(e)}")
        error_msg(f"Error during password reset: {e}")
    
    colored_input("Press Enter to continue...", Colors.INFO)

def create_sample_excel():
    """Create a sample Excel file for import."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    
    title_msg("Create Sample Excel File")
    info_msg("This will create a sample Excel file with the correct format for importing bills.")
    
    # Get filename
    default_filename = "sample_bills_template.xlsx"
    filename = colored_input(f"{Colors.PROMPT}Enter filename [{default_filename}]: {Colors.RESET}").strip()
    if not filename:
        filename = default_filename
    if not filename.lower().endswith('.xlsx'):
        filename += '.xlsx'
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Bills Template"
        
        # Define headers (same as CSV)
        headers = [
            'name', 'due_date', 'billing_cycle', 'reminder_days', 'web_page', 
            'login_info', 'password', 'company_email', 'support_phone', 
            'billing_phone', 'customer_service_hours', 'account_number', 
            'reference_id', 'support_chat_url', 'mobile_app'
        ]
        
        # Write headers
        ws.append(headers)
        
        # Add sample data
        sample_bills = [
            {
                'name': 'Electric Bill',
                'due_date': '2024-12-15',
                'billing_cycle': 'monthly',
                'reminder_days': 7,
                'web_page': 'https://www.electriccompany.com',
                'login_info': 'username@email.com',
                'password': 'your_password_here',
                'company_email': 'support@electriccompany.com',
                'support_phone': '1-800-ELECTRIC',
                'billing_phone': '1-800-BILLING',
                'customer_service_hours': 'Mon-Fri 8AM-6PM',
                'account_number': '123456789',
                'reference_id': 'ELEC-2024-001',
                'support_chat_url': 'https://www.electriccompany.com/chat',
                'mobile_app': 'Electric Company App'
            },
            {
                'name': 'Internet Service',
                'due_date': '2024-12-20',
                'billing_cycle': 'monthly',
                'reminder_days': 5,
                'web_page': 'https://www.internetprovider.com',
                'login_info': 'your_username',
                'password': 'your_password_here',
                'company_email': 'help@internetprovider.com',
                'support_phone': '1-800-INTERNET',
                'billing_phone': '1-800-BILLING',
                'customer_service_hours': '24/7',
                'account_number': '987654321',
                'reference_id': 'INT-2024-002',
                'support_chat_url': 'https://www.internetprovider.com/support',
                'mobile_app': 'Internet Provider App'
            },
            {
                'name': 'Phone Bill',
                'due_date': '2024-12-25',
                'billing_cycle': 'monthly',
                'reminder_days': 3,
                'web_page': 'https://www.phonecompany.com',
                'login_info': 'phone_number@phonecompany.com',
                'password': 'your_password_here',
                'company_email': 'support@phonecompany.com',
                'support_phone': '1-800-PHONE',
                'billing_phone': '1-800-BILLING',
                'customer_service_hours': 'Mon-Sun 7AM-10PM',
                'account_number': '555-123-4567',
                'reference_id': 'PHONE-2024-003',
                'support_chat_url': 'https://www.phonecompany.com/help',
                'mobile_app': 'Phone Company App'
            }
        ]
        
        # Write sample data
        for bill in sample_bills:
            row = [bill.get(header, '') for header in headers]
            ws.append(row)
        
        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max(12, min(max_length + 2, 40))
        
        wb.save(filename)
        success_msg(f"Sample Excel file created: {filename}")
        info_msg(f"File location: {os.path.abspath(filename)}")
        print(f"\n{Colors.INFO}📋 Sample file includes:{Colors.RESET}")
        print("  • Correct column headers")
        print("  • Example data for all fields")
        print("  • Proper date format (YYYY-MM-DD)")
        print("  • Contact information examples")
        print("  • Website URLs and login info")
        
    except Exception as e:
        error_msg(f"Error creating sample Excel file: {str(e)}")
    
    colored_input("Press Enter to continue...", Colors.INFO)

def set_master_password():
    """Set up the master password for the first time."""
    if os.path.exists(MASTER_PASSWORD_FILE):
        info_msg("Master password already set.")
        return None
    
    print(f"\n{Colors.TITLE}🔐 Set Master Password{Colors.RESET}")
    print("This password will protect all your bill passwords.")
    print("Make sure to remember it - it cannot be recovered if forgotten!")
    
    while True:
        password = getpass.getpass("Enter master password: ").strip()
        if len(password) < 8:
            error_msg("Password must be at least 8 characters long.")
            continue
        
        confirm_password = getpass.getpass("Confirm master password: ").strip()
        if password != confirm_password:
            error_msg("Passwords do not match. Please try again.")
            continue
        
        # Hash and save the password
        salt = os.urandom(16)
        password_hash = hashlib.pbkdf2_hmac('sha256', password.encode(), salt, 100000)
        
        with open(MASTER_PASSWORD_FILE, 'wb') as f:
            f.write(salt + password_hash)
        
        success_msg("Master password set successfully!")
        return password

def verify_master_password():
    """Verify the master password and return it if correct."""
    if not os.path.exists(MASTER_PASSWORD_FILE):
        # First time setup
        return set_master_password()
    
    print(f"\n{Colors.TITLE}🔐 Master Password Required{Colors.RESET}")
    
    while True:
        password = getpass.getpass("Enter master password: ").strip()
        
        # Read stored password hash
        with open(MASTER_PASSWORD_FILE, 'rb') as f:
            data = f.read()
            salt = data[:16]
            stored_hash = data[16:]
        
        # Verify password
        password_hash = hashlib.pbkdf2_hmac('sha256', password.encode(), salt, 100000)
        
        if password_hash == stored_hash:
            success_msg("Password verified successfully!")
            return password
        else:
            error_msg("Incorrect password. Please try again.")

def show_password_recovery_options():
    """Show password recovery options and guidance."""
    clear_console()
    title_msg("Password Recovery Options")
    
    print(f"{Colors.TITLE}🔑 Master Password Recovery{Colors.RESET}")
    print("If you've forgotten your master password, here are your options:")
    print()
    
    print(f"{Colors.WARNING}Option 1: Reset Master Password{Colors.RESET}")
    print("• This will remove the current master password")
    print("• You'll need to set a new master password")
    print("• Encrypted bill passwords may need to be re-entered")
    print("• A backup of your current data will be created")
    print()
    
    print(f"{Colors.WARNING}Option 2: Export Bills for Recovery{Colors.RESET}")
    print("• Export your bills with decrypted passwords")
    print("• Save the export file in a secure location")
    print("• Use this file to restore your bills if needed")
    print()
    
    print(f"{Colors.WARNING}Option 3: Manual Recovery{Colors.RESET}")
    print("• Check for backup files in the backups directory")
    print("• Look for password reset backup directories")
    print("• Restore from a previous backup if available")
    print()
    
    print(f"{Colors.TITLE}💡 Prevention Tips{Colors.RESET}")
    print("• Write down your master password in a secure location")
    print("• Use a password manager for additional security")
    print("• Create regular backups of your data")
    print("• Export bills periodically for safekeeping")
    print()
    
    print(f"{Colors.TITLE}🚨 Important Notes{Colors.RESET}")
    print("• Master passwords cannot be recovered if forgotten")
    print("• Encrypted data may be lost if password is forgotten")
    print("• Always keep backups in multiple secure locations")
    print("• Consider using a password manager for the master password")
    print()
    
    colored_input("Press Enter to continue...", Colors.INFO)

def export_bills_for_recovery():
    """Export bills with decrypted passwords for recovery purposes."""
    title_msg("Export Bills for Recovery")
    info_msg("This will export your bills with decrypted passwords for backup/recovery.")
    warning_msg("⚠️  WARNING: This file will contain sensitive information in plain text!")
    print()
    
    confirm = colored_input("Are you sure you want to export bills with decrypted passwords? (yes/no): ", Colors.WARNING).strip().lower()
    if confirm not in ['yes', 'y']:
        info_msg("Export cancelled.")
        colored_input("Press Enter to continue...", Colors.INFO)
        return
    
    # Create recovery export
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    recovery_file = f"bills_recovery_export_{timestamp}.json"
    
    try:
        # Create a copy of bills with decrypted passwords
        recovery_bills = []
        for bill in bills:
            recovery_bill = bill.copy()
            
            # Decrypt password if available
            if 'password' in recovery_bill and recovery_bill['password']:
                if CRYPTOGRAPHY_AVAILABLE and password_encryption.fernet:
                    decrypted_password = password_encryption.decrypt_password(recovery_bill['password'])
                    recovery_bill['password'] = decrypted_password
                    recovery_bill['password_encrypted'] = True
                else:
                    recovery_bill['password_encrypted'] = False
            
            recovery_bills.append(recovery_bill)
        
        # Add metadata
        recovery_data = {
            'export_date': datetime.now().isoformat(),
            'export_type': 'recovery_export',
            'total_bills': len(recovery_bills),
            'encryption_available': CRYPTOGRAPHY_AVAILABLE,
            'bills': recovery_bills
        }
        
        # Save to file
        with open(recovery_file, 'w') as f:
            json.dump(recovery_data, f, indent=2)
        
        success_msg(f"Recovery export created: {recovery_file}")
        warning_msg("⚠️  Keep this file secure - it contains sensitive information!")
        info_msg("Store this file in a safe location for emergency recovery.")
        
    except Exception as e:
        error_msg(f"Error creating recovery export: {e}")
    
    colored_input("Press Enter to continue...", Colors.INFO)

def view_backup_files():
    """View available backup files."""
    clear_console()
    title_msg("Backup Files")
    
    backup_files = []
    
    # Check for bills backups
    if os.path.exists(BACKUP_DIR):
        for file in os.listdir(BACKUP_DIR):
            if file.endswith('.json'):
                file_path = os.path.join(BACKUP_DIR, file)
                file_size = os.path.getsize(file_path)
                file_time = datetime.fromtimestamp(os.path.getmtime(file_path))
                backup_files.append({
                    'name': file,
                    'path': file_path,
                    'size': file_size,
                    'time': file_time,
                    'type': 'bills_backup'
                })
    
    # Check for password reset backups
    for item in os.listdir('.'):
        if item.startswith('password_reset_backup_') and os.path.isdir(item):
            backup_files.append({
                'name': item,
                'path': item,
                'size': 'Directory',
                'time': datetime.fromtimestamp(os.path.getmtime(item)),
                'type': 'password_reset_backup'
            })
    
    if not backup_files:
        info_msg("No backup files found.")
        colored_input("Press Enter to continue...", Colors.INFO)
        return
    
    # Sort by time (newest first)
    backup_files.sort(key=lambda x: x['time'], reverse=True)
    
    print(f"{Colors.INFO}Available backup files:{Colors.RESET}")
    print()
    
    for i, backup in enumerate(backup_files, 1):
        print(f"{Colors.MENU}{i:2}.{Colors.RESET} {Colors.TITLE}{backup['name']}{Colors.RESET}")
        print(f"    Type: {Colors.INFO}{backup['type']}{Colors.RESET}")
        print(f"    Date: {Colors.INFO}{backup['time'].strftime('%Y-%m-%d %H:%M:%S')}{Colors.RESET}")
        if backup['size'] != 'Directory':
            print(f"    Size: {Colors.INFO}{backup['size']} bytes{Colors.RESET}")
        print()
    
    print(f"{Colors.MENU}Options:{Colors.RESET}")
    print("1. View backup contents")
    print("2. Restore from backup")
    print("3. Delete backup")
    print("4. Back to menu")
    
    choice = colored_input("Choose option (1-4): ", Colors.PROMPT).strip()
    
    if choice == '1':
        view_backup_contents(backup_files)
    elif choice == '2':
        restore_from_backup(backup_files)
    elif choice == '3':
        delete_backup(backup_files)
    elif choice == '4':
        return
    else:
        error_msg("Invalid option.")

def view_backup_contents(backup_files):
    """View the contents of a backup file."""
    if not backup_files:
        return
    
    try:
        choice = int(colored_input(f"Enter backup number to view (1-{len(backup_files)}): ", Colors.PROMPT))
        if 1 <= choice <= len(backup_files):
            backup = backup_files[choice - 1]
            
            if backup['type'] == 'bills_backup':
                try:
                    with open(backup['path'], 'r') as f:
                        backup_data = json.load(f)
                    
                    print(f"\n{Colors.TITLE}Backup Contents: {backup['name']}{Colors.RESET}")
                    print(f"Total bills: {len(backup_data)}")
                    print()
                    
                    for i, bill in enumerate(backup_data[:5], 1):  # Show first 5 bills
                        print(f"{i}. {bill.get('name', 'Unknown')} - {bill.get('due_date', 'No date')}")
                    
                    if len(backup_data) > 5:
                        print(f"... and {len(backup_data) - 5} more bills")
                    
                except Exception as e:
                    error_msg(f"Error reading backup: {e}")
            
            elif backup['type'] == 'password_reset_backup':
                print(f"\n{Colors.TITLE}Password Reset Backup: {backup['name']}{Colors.RESET}")
                print("This is a directory containing backup files.")
                try:
                    files = os.listdir(backup['path'])
                    print(f"Files in backup: {', '.join(files)}")
                except Exception as e:
                    error_msg(f"Error reading backup directory: {e}")
        
        else:
            error_msg("Invalid backup number.")
    except ValueError:
        error_msg("Please enter a valid number.")
    
    colored_input("Press Enter to continue...", Colors.INFO)

def restore_from_backup(backup_files):
    """Restore data from a backup file."""
    global bills
    
    if not backup_files:
        return
    
    try:
        choice = int(colored_input(f"Enter backup number to restore (1-{len(backup_files)}): ", Colors.PROMPT))
        if 1 <= choice <= len(backup_files):
            backup = backup_files[choice - 1]
            
            warning_msg(f"⚠️  This will overwrite your current data with backup: {backup['name']}")
            confirm = colored_input("Are you sure? (yes/no): ", Colors.WARNING).strip().lower()
            
            if confirm not in ['yes', 'y']:
                info_msg("Restore cancelled.")
                colored_input("Press Enter to continue...", Colors.INFO)
                return
            
            if backup['type'] == 'bills_backup':
                # Restore bills backup
                try:
                    with open(backup['path'], 'r') as f:
                        backup_data = json.load(f)
                    
                    # Create backup of current data first
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    current_backup = f"pre_restore_backup_{timestamp}.json"
                    with open(current_backup, 'w') as f:
                        json.dump(bills, f, indent=2)
                    
                    # Restore the backup data
                    bills = backup_data
                    save_bills()
                    
                    success_msg(f"Successfully restored {len(bills)} bills from backup")
                    info_msg(f"Current data backed up to: {current_backup}")
                    
                except Exception as e:
                    error_msg(f"Error restoring backup: {e}")
            
            elif backup['type'] == 'password_reset_backup':
                # Restore password reset backup
                try:
                    # This would require more complex logic to restore specific files
                    warning_msg("Password reset backup restoration requires manual file copying.")
                    info_msg(f"Backup location: {backup['path']}")
                    info_msg("You may need to manually copy files from this directory.")
                    
                except Exception as e:
                    error_msg(f"Error with password reset backup: {e}")
        
        else:
            error_msg("Invalid backup number.")
    except ValueError:
        error_msg("Please enter a valid number.")
    
    colored_input("Press Enter to continue...", Colors.INFO)

def delete_backup(backup_files):
    """Delete a backup file."""
    if not backup_files:
        return
    
    try:
        choice = int(colored_input(f"Enter backup number to delete (1-{len(backup_files)}): ", Colors.PROMPT))
        if 1 <= choice <= len(backup_files):
            backup = backup_files[choice - 1]
            
            warning_msg(f"⚠️  This will permanently delete backup: {backup['name']}")
            confirm = colored_input("Are you sure? (yes/no): ", Colors.WARNING).strip().lower()
            
            if confirm not in ['yes', 'y']:
                info_msg("Deletion cancelled.")
                colored_input("Press Enter to continue...", Colors.INFO)
                return
            
            try:
                if backup['type'] == 'password_reset_backup':
                    shutil.rmtree(backup['path'])
                else:
                    os.remove(backup['path'])
                
                success_msg(f"Successfully deleted backup: {backup['name']}")
                
            except Exception as e:
                error_msg(f"Error deleting backup: {e}")
        
        else:
            error_msg("Invalid backup number.")
    except ValueError:
        error_msg("Please enter a valid number.")
    
    colored_input("Press Enter to continue...", Colors.INFO)

def change_master_password():
    """Change the master password."""
    print(f"\n{Colors.TITLE}🔐 Change Master Password{Colors.RESET}")
    
    # Verify current password first
    current_password = getpass.getpass("Enter current master password: ").strip()
    
    # Verify current password
    with open(MASTER_PASSWORD_FILE, 'rb') as f:
        data = f.read()
        salt = data[:16]
        stored_hash = data[16:]
    
    current_hash = hashlib.pbkdf2_hmac('sha256', current_password.encode(), salt, 100000)
    
    if current_hash != stored_hash:
        error_msg("Current password is incorrect.")
        colored_input("Press Enter to continue...", Colors.INFO)
        return
    
    # Get new password
    while True:
        new_password = getpass.getpass("Enter new master password: ").strip()
        if len(new_password) < 8:
            error_msg("Password must be at least 8 characters long.")
            continue
        
        confirm_password = getpass.getpass("Confirm new master password: ").strip()
        if new_password != confirm_password:
            error_msg("Passwords do not match. Please try again.")
            continue
        
        break
    
    # Create backup before changing
    backup_timestamp = str(int(time.time()))
    backup_file = f".master_password.backup.{backup_timestamp}"
    shutil.copy2(MASTER_PASSWORD_FILE, backup_file)
    
    # Hash and save new password
    new_salt = os.urandom(16)
    new_password_hash = hashlib.pbkdf2_hmac('sha256', new_password.encode(), new_salt, 100000)
    
    with open(MASTER_PASSWORD_FILE, 'wb') as f:
        f.write(new_salt + new_password_hash)
    
    # Re-encrypt all passwords with new master password
    if CRYPTOGRAPHY_AVAILABLE:
        re_encrypt_passwords_with_new_master(current_password, new_password)
    
    success_msg("Master password changed successfully!")
    info_msg(f"Backup created: {backup_file}")
    colored_input("Press Enter to continue...", Colors.INFO)

def reset_master_password():
    """Reset the master password (destructive operation)."""
    print(f"\n{Colors.TITLE}🔄 Reset Master Password{Colors.RESET}")
    warning_msg("⚠️  WARNING: This will remove the current master password!")
    warning_msg("⚠️  Encrypted bill passwords may become inaccessible!")
    print()
    
    confirm = colored_input("Are you sure you want to reset the master password? (yes/no): ", Colors.WARNING).strip().lower()
    if confirm not in ['yes', 'y']:
        info_msg("Password reset cancelled.")
        colored_input("Press Enter to continue...", Colors.INFO)
        return
    
    # Create comprehensive backup
    backup_timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_dir = f"password_reset_backup_{backup_timestamp}"
    os.makedirs(backup_dir, exist_ok=True)
    
    # Backup all important files
    files_to_backup = [BILLS_FILE, TEMPLATES_FILE, MASTER_PASSWORD_FILE]
    if os.path.exists(ENCRYPTION_KEY_FILE):
        files_to_backup.append(ENCRYPTION_KEY_FILE)
    if os.path.exists(SALT_FILE):
        files_to_backup.append(SALT_FILE)
    
    for file_path in files_to_backup:
        if os.path.exists(file_path):
            shutil.copy2(file_path, os.path.join(backup_dir, os.path.basename(file_path)))
    
    # Remove password files
    if os.path.exists(MASTER_PASSWORD_FILE):
        os.remove(MASTER_PASSWORD_FILE)
    if os.path.exists(ENCRYPTION_KEY_FILE):
        os.remove(ENCRYPTION_KEY_FILE)
    if os.path.exists(SALT_FILE):
        os.remove(SALT_FILE)
    
    success_msg("Master password reset successfully!")
    info_msg(f"Backup created in: {backup_dir}")
    warning_msg("You will need to set a new master password on next startup.")
    colored_input("Press Enter to continue...", Colors.INFO)

def re_encrypt_passwords_with_new_master(old_password, new_password):
    """Re-encrypt all bill passwords with the new master password."""
    if not CRYPTOGRAPHY_AVAILABLE:
        return
    
    try:
        # Initialize encryption with old password
        old_encryption = PasswordEncryption()
        old_encryption.initialize_encryption(old_password)
        
        # Initialize encryption with new password
        new_encryption = PasswordEncryption()
        new_encryption.initialize_encryption(new_password)
        
        # Re-encrypt all bill passwords
        for bill in bills:
            if bill.get('password'):
                # Decrypt with old password
                decrypted = old_encryption.decrypt_password(bill['password'])
                # Encrypt with new password
                bill['password'] = new_encryption.encrypt_password(decrypted)
        
        # Save bills with re-encrypted passwords
        save_bills()
        
        success_msg("All passwords re-encrypted with new master password.")
        
    except Exception as e:
        error_msg(f"Error re-encrypting passwords: {e}")

def export_bills_for_recovery():
    """Export bills with decrypted passwords for recovery purposes."""
    title_msg("Export Bills for Recovery")
    info_msg("This will export your bills with decrypted passwords for backup/recovery.")
    warning_msg("⚠️  WARNING: This file will contain sensitive information in plain text!")
    print()
    
    confirm = colored_input("Are you sure you want to export bills with decrypted passwords? (yes/no): ", Colors.WARNING).strip().lower()
    if confirm not in ['yes', 'y']:
        info_msg("Export cancelled.")
        colored_input("Press Enter to continue...", Colors.INFO)
        return
    
    # Create recovery export
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    recovery_file = f"bills_recovery_export_{timestamp}.json"
    
    try:
        # Create a copy of bills with decrypted passwords
        recovery_bills = []
        for bill in bills:
            recovery_bill = bill.copy()
            
            # Decrypt password if available
            if 'password' in recovery_bill and recovery_bill['password']:
                if CRYPTOGRAPHY_AVAILABLE and password_encryption.fernet:
                    decrypted_password = password_encryption.decrypt_password(recovery_bill['password'])
                    recovery_bill['password'] = decrypted_password
                    recovery_bill['password_encrypted'] = True
                else:
                    recovery_bill['password_encrypted'] = False
            
            recovery_bills.append(recovery_bill)
        
        # Add metadata
        recovery_data = {
            'export_date': datetime.now().isoformat(),
            'export_type': 'recovery_export',
            'total_bills': len(recovery_bills),
            'encryption_available': CRYPTOGRAPHY_AVAILABLE,
            'bills': recovery_bills
        }
        
        # Save to file
        with open(recovery_file, 'w') as f:
            json.dump(recovery_data, f, indent=2)
        
        success_msg(f"Recovery export created: {recovery_file}")
        warning_msg("⚠️  Keep this file secure - it contains sensitive information!")
        info_msg("Store this file in a safe location for emergency recovery.")
        
    except Exception as e:
        error_msg(f"Error creating recovery export: {e}")
    
    colored_input("Press Enter to continue...", Colors.INFO)

def password_management_menu():
    """Display password management menu."""
    while True:
        clear_console()
        title_msg("Password Management")
        
        print(f"{Colors.MENU}1.{Colors.RESET} 🔑 Change master password")
        print(f"{Colors.MENU}2.{Colors.RESET} 🔄 Reset master password")
        print(f"{Colors.MENU}3.{Colors.RESET} 📋 Password recovery options")
        print(f"{Colors.MENU}4.{Colors.RESET} 📤 Export bills for recovery")
        print(f"{Colors.MENU}5.{Colors.RESET} 📁 View backup files")
        print(f"{Colors.MENU}6.{Colors.RESET} 🚪 Back to main menu")
        
        choice = colored_input("\nChoose option (1-6): ", Colors.PROMPT).strip()
        
        if choice == '1':
            clear_console()
            change_master_password()
        elif choice == '2':
            clear_console()
            reset_master_password()
        elif choice == '3':
            clear_console()
            show_password_recovery_options()
        elif choice == '4':
            clear_console()
            export_bills_for_recovery()
        elif choice == '5':
            clear_console()
            view_backup_files()
        elif choice == '6':
            break
        else:
            error_msg("Invalid option. Please choose 1-6.")
            colored_input("Press Enter to continue...", Colors.WARNING)

# ... existing code ...

# Entry point
if __name__ == "__main__":
    main()